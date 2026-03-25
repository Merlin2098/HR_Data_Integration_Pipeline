from __future__ import annotations

from datetime import date
from pathlib import Path

import polars as pl

from src.modules.control_practicantes.steps.step1_controlpracticantes import validar_esquema
from src.modules.control_practicantes.steps.step2_controlpracticantes import generar_gold_con_flags
from src.utils.bd_document_date import (
    BD_DOCUMENT_DATE_COLUMN,
    append_control_practicantes_document_date_column,
    extract_control_practicantes_document_date,
)
from src.utils.validate_source import load_validation_contract, validate_excel_source


def test_extract_control_practicantes_document_date_accepts_supported_extensions():
    assert extract_control_practicantes_document_date(
        Path("BD Practicantes 30.03.2026.xlsx")
    ) == "2026-03-30"
    assert extract_control_practicantes_document_date(
        Path("BD Practicantes 30.03.2026.xlsm")
    ) == "2026-03-30"
    assert extract_control_practicantes_document_date(
        Path("BD Practicantes 30.03.2026.xls")
    ) == "2026-03-30"


def test_extract_control_practicantes_document_date_rejects_invalid_pattern():
    try:
        extract_control_practicantes_document_date(
            Path("Control Practicantes 30.03.2026.xlsx")
        )
    except ValueError as exc:
        assert "BD Practicantes DD.MM.YYYY.xlsx|xlsm|xls" in str(exc)
    else:
        raise AssertionError("Expected ValueError for invalid filename pattern")


def test_extract_control_practicantes_document_date_rejects_invalid_calendar_date():
    try:
        extract_control_practicantes_document_date(Path("BD Practicantes 31.02.2026.xlsx"))
    except ValueError as exc:
        assert "Invalid document date" in str(exc)
    else:
        raise AssertionError("Expected ValueError for invalid filename date")


def test_control_practicantes_step1_adds_fecha_documento_after_schema_validation():
    df = pl.DataFrame(
        {
            "N°": ["1"],
            "DNI": ["12345678"],
            "APELLIDOS Y NOMBRES": ["Ana Perez"],
            "CONDICION": ["PRACTICANTE PROFESIONAL"],
            "FECHA ING": [date(2025, 4, 1)],
            "F. RENOVACION": [date(2026, 3, 30)],
            "SEDE": ["Lima"],
            "UNIVERSIDAD": ["UNI"],
            "JEFE INMEDIATO": ["Jefe"],
            "GERENCIA": ["RRHH"],
        }
    )
    esquema = {
        "required_columns": [
            "N°",
            "DNI",
            "APELLIDOS Y NOMBRES",
            "CONDICION",
            "FECHA ING",
            "F. RENOVACION",
            "SEDE",
            "UNIVERSIDAD",
            "JEFE INMEDIATO",
            "GERENCIA",
        ],
        "column_types": {
            "N°": "string",
            "DNI": "string",
            "APELLIDOS Y NOMBRES": "string",
            "CONDICION": "string",
            "FECHA ING": "date",
            "F. RENOVACION": "date",
            "SEDE": "string",
            "UNIVERSIDAD": "string",
            "JEFE INMEDIATO": "string",
            "GERENCIA": "string",
        },
    }

    es_valido, errores = validar_esquema(df, esquema)
    assert es_valido, errores

    result = append_control_practicantes_document_date_column(
        df,
        Path("BD Practicantes 30.03.2026.xlsx"),
    )

    assert result.columns[-1] == BD_DOCUMENT_DATE_COLUMN
    assert result[BD_DOCUMENT_DATE_COLUMN].to_list() == ["2026-03-30"]


def test_control_practicantes_gold_query_keeps_fecha_documento(tmp_path: Path):
    silver_path = tmp_path / "control_practicantes_silver.parquet"
    df_silver = pl.DataFrame(
        {
            "DNI": ["12345678"],
            "CONDICION": ["PRACTICANTE PROFESIONAL"],
            "FECHA ING": [pl.Series("FECHA ING", ["2025-04-01"]).str.to_date("%Y-%m-%d")[0]],
            "F. RENOVACION": [pl.Series("F. RENOVACION", ["2026-03-30"]).str.to_date("%Y-%m-%d")[0]],
            "SEDE": ["Lima"],
            "UNIVERSIDAD": ["UNI"],
            "JEFE INMEDIATO": ["Jefe"],
            "GERENCIA": ["RRHH"],
            "FECHA_DOCUMENTO": ["2026-03-30"],
        }
    )
    df_silver.write_parquet(silver_path)

    result = generar_gold_con_flags(silver_path)

    assert "FECHA_DOCUMENTO" in result.columns
    assert result["FECHA_DOCUMENTO"].to_list() == ["2026-03-30"]


def test_control_practicantes_preflight_rejects_invalid_filename_date(tmp_path: Path):
    workbook = tmp_path / "BD Practicantes 31.02.2026.xlsx"

    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Practicantes"
    headers = [
        "N°",
        "DNI",
        "APELLIDOS Y NOMBRES",
        "CONDICION",
        "FECHA ING",
        "F. RENOVACION",
        "SEDE",
        "UNIVERSIDAD",
        "JEFE INMEDIATO",
        "GERENCIA",
    ]
    for idx, header in enumerate(headers, start=1):
        ws.cell(row=4, column=idx).value = header
    ws["A5"] = "1"
    ws["B5"] = "12345678"
    wb.save(workbook)
    wb.close()

    contract = load_validation_contract("control_practicantes")
    report = validate_excel_source(workbook, contract)

    assert not report.passed
    assert any("Invalid document date in filename" in error for error in report.errors)
