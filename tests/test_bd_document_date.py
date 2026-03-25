from __future__ import annotations

from datetime import date
from pathlib import Path

import polars as pl

from src.modules.bd.steps.step1_capasilver import crear_dataframe_polars
from src.modules.bd.steps.step2_capagold import (
    convertir_columnas_fecha,
    dividir_por_modalidad,
    validar_y_filtrar_columnas,
)
from src.modules.bd.steps.step3_flags_empleados import aplicar_flags_duckdb
from src.utils.bd_document_date import (
    BD_DOCUMENT_DATE_COLUMN,
    append_bd_document_date_column,
    extract_bd_document_date,
)
from src.utils.validate_source import load_validation_contract, validate_excel_source


def test_extract_bd_document_date_accepts_supported_extensions():
    assert extract_bd_document_date(Path("BD. 21.12.2025..xlsx")) == "2025-12-21"
    assert extract_bd_document_date(Path("BD. 21.12.2025..xlsm")) == "2025-12-21"
    assert extract_bd_document_date(Path("BD. 21.12.2025..xls")) == "2025-12-21"


def test_extract_bd_document_date_rejects_invalid_pattern():
    try:
        extract_bd_document_date(Path("Base de Datos 21.12.2025.xlsm"))
    except ValueError as exc:
        assert "BD. DD.MM.YYYY..xlsx|xlsm|xls" in str(exc)
    else:
        raise AssertionError("Expected ValueError for invalid filename pattern")


def test_extract_bd_document_date_rejects_invalid_calendar_date():
    try:
        extract_bd_document_date(Path("BD. 31.02.2025..xlsm"))
    except ValueError as exc:
        assert "Invalid document date" in str(exc)
    else:
        raise AssertionError("Expected ValueError for invalid filename date")


def test_step1_adds_fecha_documento_constant_column():
    headers = ["NUMERO DE DOC", "NOMBRE COMPLETOS"]
    data_rows = [["123", "Ana"], ["456", "Luis"]]

    df = crear_dataframe_polars(headers, data_rows)
    df = append_bd_document_date_column(df, Path("BD. 21.12.2025..xlsm"))

    assert df.columns == ["NUMERO DE DOC", "NOMBRE COMPLETOS", BD_DOCUMENT_DATE_COLUMN]
    assert df[BD_DOCUMENT_DATE_COLUMN].to_list() == ["2025-12-21", "2025-12-21"]


def test_step2_preserves_fecha_documento_in_gold_outputs():
    df_silver = pl.DataFrame(
        {
            "NUMERO DE DOC": ["1", "2"],
            "CODIGO SAP2": ["A1", "A2"],
            "NOMBRE COMPLETOS": ["Ana", "Luis"],
            "GERENCIA": ["G1", "G2"],
            "SEXO": ["F", "M"],
            "SEDE3": ["Lima", "Arequipa"],
            "WHITE COLLAR / BLUE COLLAR": ["WHITE", "BLUE"],
            "Modalidad de Contrato": [
                "CONTRATO INDETERMINADO",
                "TERMINO DE CONVENIO PRACTICAS",
            ],
            "Fecha de Termino": ["2025-12-31 00:00:00", "2025-06-30 00:00:00"],
            "SERVICIO": ["Ops", "Ops"],
            "REGIMEN DE TRABAJO": ["5x2", "5x2"],
            "FECH_INGR.": ["2020-01-01 00:00:00", "2024-01-01 00:00:00"],
            "FECHA DE NAC.": ["1990-05-20 00:00:00", "2000-08-15 00:00:00"],
            "FECHA_DOCUMENTO": ["2025-12-21", "2025-12-21"],
        }
    )
    esquema = {
        "columns": [
            {"name": "NUMERO DE DOC", "type": "string", "required": True},
            {"name": "CODIGO SAP2", "type": "string", "required": True},
            {"name": "NOMBRE COMPLETOS", "type": "string", "required": True},
            {"name": "GERENCIA", "type": "string", "required": True},
            {"name": "SEXO", "type": "string", "required": True},
            {"name": "SEDE3", "type": "string", "required": True},
            {"name": "WHITE COLLAR / BLUE COLLAR", "type": "string", "required": True},
            {"name": "Modalidad de Contrato", "type": "string", "required": True},
            {"name": "Fecha de Termino", "type": "date", "required": True},
            {"name": "SERVICIO", "type": "string", "required": True},
            {"name": "REGIMEN DE TRABAJO", "type": "string", "required": True},
            {"name": "FECH_INGR.", "type": "date", "required": True},
            {"name": "FECHA DE NAC.", "type": "date", "required": True},
            {"name": "FECHA_DOCUMENTO", "type": "string", "required": True},
        ]
    }

    df_gold, stats = validar_y_filtrar_columnas(df_silver, esquema)
    df_gold = convertir_columnas_fecha(df_gold, esquema)
    df_empleados, df_practicantes = dividir_por_modalidad(df_gold)

    assert stats["missing_list"] == []
    assert "FECHA_DOCUMENTO" in df_empleados.columns
    assert "FECHA_DOCUMENTO" in df_practicantes.columns
    assert df_empleados["FECHA_DOCUMENTO"].to_list() == ["2025-12-21"]
    assert df_practicantes["FECHA_DOCUMENTO"].to_list() == ["2025-12-21"]


def test_flags_query_keeps_fecha_documento_from_gold():
    df = pl.DataFrame(
        {
            "NUMERO DE DOC": ["1"],
            "FECHA DE NAC.": [date(1960, 1, 1)],
            "FECH_INGR.": [date(2020, 1, 1)],
            "Fecha de Termino": [date(2025, 12, 31)],
            "Modalidad de Contrato": ["CONTRATO INDETERMINADO"],
            "FECHA_DOCUMENTO": ["2025-12-21"],
        }
    )
    query = Path("assets/queries/queries_flags_gold.sql").read_text(encoding="utf-8")

    result = aplicar_flags_duckdb(df, query)

    assert "FECHA_DOCUMENTO" in result.columns
    assert result["FECHA_DOCUMENTO"].to_list() == ["2025-12-21"]


def test_bd_preflight_rejects_invalid_filename_date(tmp_path: Path):
    workbook = tmp_path / "BD. 31.02.2025..xlsm"

    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "METSO"
    ws["A10"] = "NUMERO DE DOC"
    ws["B10"] = "CODIGO SAP2"
    ws["C10"] = "NOMBRE COMPLETOS"
    ws["D10"] = "GERENCIA"
    ws["E10"] = "SEXO"
    ws["F10"] = "SEDE3"
    ws["G10"] = "WHITE COLLAR / BLUE COLLAR"
    ws["H10"] = "Modalidad de Contrato"
    ws["I10"] = "Fecha de Termino"
    ws["J10"] = "SERVICIO"
    ws["K10"] = "REGIMEN DE TRABAJO"
    ws["L10"] = "FECH_INGR."
    ws["M10"] = "FECHA DE NAC."
    ws["A11"] = "123"
    wb.save(workbook)
    wb.close()

    contract = load_validation_contract("bd")
    report = validate_excel_source(workbook, contract)

    assert not report.passed
    assert any("Invalid document date in filename" in error for error in report.errors)
