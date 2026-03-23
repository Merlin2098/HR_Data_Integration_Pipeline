from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
import re
from typing import Any

from src.utils.structured_config import load_structured_data, resolve_structured_path


EXCEL_EXTENSIONS = {".xlsx", ".xlsm", ".xls"}


class SourceValidationError(ValueError):
    """Raised when source contract validation fails."""

    def __init__(self, report: "ValidationReport"):
        super().__init__(report.error_summary())
        self.report = report


@dataclass
class ValidationReport:
    contract_id: str
    contract_name: str
    checked_sources: list[Path] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)

    @property
    def passed(self) -> bool:
        return not self.errors

    def merge(self, other: "ValidationReport") -> None:
        self.checked_sources.extend(other.checked_sources)
        self.errors.extend(other.errors)

    def add_error(self, source: Path, message: str) -> None:
        self.errors.append(f"{source.name}: {message}")

    def error_summary(self) -> str:
        lines = [f"Preflight validation failed for '{self.contract_id}' ({self.contract_name})"]
        lines.extend(f"- {error}" for error in self.errors)
        return "\n".join(lines)

    def raise_if_failed(self) -> None:
        if not self.passed:
            raise SourceValidationError(self)


def load_validation_contract(etl_name: str) -> dict[str, Any]:
    """Loads a source validation contract by ETL/contract id."""
    path = resolve_structured_path(f"assets/validate_source/{etl_name}")
    if not path.exists():
        raise FileNotFoundError(f"Validation contract not found: {path}")
    contract = load_structured_data(path)

    if "id" not in contract:
        contract["id"] = etl_name
    if "name" not in contract:
        contract["name"] = etl_name

    return contract


def validate_excel_source(
    path: Path, contract: dict[str, Any], required_cols: list[str] | None = None
) -> ValidationReport:
    """
    Validates one Excel source file against the provided contract.
    """
    report = ValidationReport(
        contract_id=str(contract.get("id", "unknown")),
        contract_name=str(contract.get("name", contract.get("id", "unknown"))),
        checked_sources=[Path(path)],
    )

    file_path = Path(path)

    if not file_path.exists():
        report.add_error(file_path, "Source file does not exist")
        return report

    file_type = contract.get("file_type", {})
    expected_extensions = {
        ext.lower() for ext in file_type.get("expected_extensions", list(EXCEL_EXTENSIONS))
    }
    if file_path.suffix.lower() not in expected_extensions:
        report.add_error(
            file_path,
            f"Invalid file extension '{file_path.suffix}'. Expected one of {sorted(expected_extensions)}",
        )
        return report

    filename_regex = contract.get("filename_regex")
    if filename_regex and not re.search(filename_regex, file_path.name):
        report.add_error(file_path, f"Filename does not match regex: {filename_regex}")

    try:
        import openpyxl

        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    except Exception as exc:
        report.add_error(file_path, f"Unable to open Excel file: {exc}")
        return report

    try:
        sheet_rules = contract.get("sheet_rules", {})
        sheets_cfg: dict[str, Any] = contract.get("sheets", {})

        required_sheets = list(sheet_rules.get("required_sheets", sheets_cfg.keys()))
        allowed_sheets = set(sheet_rules.get("allowed_sheets", required_sheets))
        enforce_allowed_sheets = bool(sheet_rules.get("enforce_allowed_sheets", False))

        available_sheets = set(wb.sheetnames)
        for sheet_name in required_sheets:
            if sheet_name not in available_sheets:
                report.add_error(file_path, f"Missing sheet '{sheet_name}'")

        if enforce_allowed_sheets:
            unexpected = sorted(available_sheets - allowed_sheets)
            if unexpected:
                report.add_error(file_path, f"Unexpected sheet(s): {unexpected}")

        for sheet_name, cfg in sheets_cfg.items():
            if sheet_name not in available_sheets:
                continue

            ws = wb[sheet_name]
            header_start = cfg.get("header_start", {})
            row_idx = int(header_start.get("row_index_1_based", 1))
            col_letter = str(header_start.get("column_letter", "A"))
            col_idx = column_letter_to_index(col_letter)

            start_value = ws.cell(row=row_idx, column=col_idx).value
            if _is_empty(start_value):
                report.add_error(
                    file_path,
                    (
                        f"Expected header at row {row_idx} col {col_letter.upper()} "
                        f"(sheet '{sheet_name}') but found empty/shifted"
                    ),
                )

            header_values = _read_header_values(
                worksheet=ws,
                row_index=row_idx,
                start_col_index=col_idx,
                max_scan_cols=int(cfg.get("max_scan_cols", 250)),
            )

            required = (
                list(required_cols) if required_cols is not None else _resolve_required_columns(cfg)
            )
            missing_required = _missing_columns(required, header_values)
            if missing_required:
                report.add_error(
                    file_path,
                    f"Missing required columns in sheet '{sheet_name}': {missing_required}",
                )
    finally:
        wb.close()

    return report


def validate_all_sources_for_etl(
    etl_name: str, inputs: dict[str, Any] | Path | list[Path] | str
) -> ValidationReport:
    """
    Validates one or many source files for a single ETL contract id.
    """
    contract = load_validation_contract(etl_name)
    aggregate = ValidationReport(
        contract_id=str(contract.get("id", etl_name)),
        contract_name=str(contract.get("name", etl_name)),
    )

    paths = _coerce_paths(inputs)
    if not paths:
        aggregate.errors.append("No source files were provided for preflight validation")
        return aggregate

    for source in paths:
        source_report = validate_excel_source(Path(source), contract)
        aggregate.merge(source_report)

    return aggregate


def column_letter_to_index(column_letter: str) -> int:
    """Converts Excel column letter to 1-based integer index."""
    letter = str(column_letter).strip().upper()
    if not letter or not letter.isalpha():
        raise ValueError(f"Invalid Excel column letter: {column_letter!r}")

    value = 0
    for char in letter:
        value = (value * 26) + (ord(char) - ord("A") + 1)
    return value


def _coerce_paths(inputs: dict[str, Any] | Path | list[Path] | str) -> list[Path]:
    if isinstance(inputs, (str, Path)):
        return [Path(inputs)]

    if isinstance(inputs, dict):
        out: list[Path] = []
        for value in inputs.values():
            out.extend(_coerce_paths(value))
        return out

    if isinstance(inputs, list):
        return [Path(item) for item in inputs]

    raise TypeError(f"Unsupported inputs type: {type(inputs)!r}")


def _resolve_required_columns(sheet_cfg: dict[str, Any]) -> list[str]:
    source_cfg = sheet_cfg.get("required_columns_source", {})
    if isinstance(source_cfg, list):
        return [str(item).strip() for item in source_cfg if str(item).strip()]

    if not isinstance(source_cfg, dict):
        return []

    explicit = source_cfg.get("required_columns")
    if explicit:
        return [str(item).strip() for item in explicit if str(item).strip()]

    schema_path = source_cfg.get("schema_path")
    extract_mode = str(source_cfg.get("extract", "")).strip().lower()
    if not schema_path or not extract_mode:
        return []

    schema_file = resolve_structured_path(str(schema_path))
    schema_json = load_structured_data(schema_file)

    columns = _extract_columns_from_schema(schema_json, extract_mode, source_cfg)

    excluded = {
        _normalize_header_name(item)
        for item in source_cfg.get("exclude_columns", [])
        if str(item).strip()
    }
    if excluded:
        columns = [col for col in columns if _normalize_header_name(col) not in excluded]

    return columns


def _extract_columns_from_schema(
    schema_json: dict[str, Any], extract_mode: str, source_cfg: dict[str, Any]
) -> list[str]:
    if extract_mode == "schema_non_nullable":
        schema = _get_schema_section(schema_json, source_cfg)
        return [
            str(col_name)
            for col_name, col_meta in schema.items()
            if isinstance(col_meta, dict) and col_meta.get("nullable") is False
        ]

    if extract_mode == "schema_keys":
        schema = _get_schema_section(schema_json, source_cfg)
        return [str(col_name) for col_name in schema.keys()]

    if extract_mode == "columns_required_true":
        columns = schema_json.get("columns", [])
        out: list[str] = []
        for col in columns:
            if not isinstance(col, dict):
                continue
            if col.get("required") is True and col.get("name"):
                out.append(str(col["name"]))
        return out

    if extract_mode == "columns_non_nullable":
        columns = schema_json.get("columns", [])
        out = []
        for col in columns:
            if not isinstance(col, dict):
                continue
            if col.get("nullable") is False and col.get("name"):
                out.append(str(col["name"]))
        return out

    if extract_mode == "required_columns_array":
        required_columns = schema_json.get("required_columns", [])
        return [str(col) for col in required_columns]

    return []


def _get_schema_section(schema_json: dict[str, Any], source_cfg: dict[str, Any]) -> dict[str, Any]:
    sheet_name = source_cfg.get("sheet_name")
    if sheet_name:
        hojas = schema_json.get("hojas", {})
        hoja_cfg = hojas.get(str(sheet_name), {})
        schema = hoja_cfg.get("schema", {})
        if isinstance(schema, dict):
            return schema
    schema = schema_json.get("schema", {})
    if isinstance(schema, dict):
        return schema
    return {}


def _read_header_values(
    worksheet: Any,
    row_index: int,
    start_col_index: int,
    max_scan_cols: int = 250,
) -> list[str]:
    headers: list[str] = []
    empty_streak = 0
    scan_limit = start_col_index + max_scan_cols

    for col_idx in range(start_col_index, scan_limit + 1):
        value = worksheet.cell(row=row_index, column=col_idx).value
        if _is_empty(value):
            empty_streak += 1
            if headers and empty_streak >= 3:
                break
            continue

        empty_streak = 0
        headers.append(str(value).strip())

    return headers


def _missing_columns(required: list[str], actual: list[str]) -> list[str]:
    actual_norm = {_normalize_header_name(item) for item in actual}
    missing = []
    for item in required:
        if _normalize_header_name(item) not in actual_norm:
            missing.append(item)
    return missing


def _normalize_header_name(value: str) -> str:
    return re.sub(r"\s+", " ", str(value).strip()).upper()


def _is_empty(value: Any) -> bool:
    return value is None or str(value).strip() == ""
