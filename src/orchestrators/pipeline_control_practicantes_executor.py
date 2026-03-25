"""
Ejecutor de Pipeline Control de Practicantes
Lee archivo YAML y ejecuta 2 stages secuencialmente:
  1. Bronze → Silver (procesamiento Excel)
  2. Silver → Gold (flags y tiempo de servicio)

Compatible con señales Qt para integración en UI
"""

import yaml
from pathlib import Path
from typing import Dict, Any, Optional
from importlib import import_module
from PySide6.QtCore import QObject, Signal
import sys
import time
import traceback

from src.utils.bd_document_date import has_document_date_fragment


class PipelineControlPracticantesExecutor(QObject):
    """
    Ejecutor de pipeline basado en YAML para Control de Practicantes
    Emite señales Qt para integración con UILogger
    """
    
    # Señales Qt
    log_message = Signal(str, str)  # (nivel, mensaje)
    progress_update = Signal(int, str)  # (porcentaje, mensaje)
    stage_started = Signal(str, str)  # (nombre_stage, descripción)
    stage_completed = Signal(str, bool, float)  # (nombre_stage, éxito, duración)
    
    def __init__(self, yaml_path: Path, archivo: Path, output_dir: Path):
        """
        Inicializa el executor
        
        Args:
            yaml_path: Ruta al archivo YAML del pipeline
            archivo: Archivo Excel de control de practicantes
            output_dir: Directorio base de salida (carpeta del archivo)
        """
        super().__init__()
        self.yaml_path = yaml_path
        self.archivo = archivo
        self.output_dir = output_dir
        self.pipeline_config = None
        self.stages_results = {}
        self.last_stage_error: Optional[Dict[str, Any]] = None
        
    def _log(self, nivel: str, mensaje: str):
        """Emite señal de log"""
        self.log_message.emit(nivel, mensaje)
    
    def _progress(self, porcentaje: int, mensaje: str):
        """Emite señal de progreso"""
        self.progress_update.emit(porcentaje, mensaje)

    @staticmethod
    def _extract_stage_failure(result: Any) -> Optional[str]:
        """
        Detecta fallas cuando un step retorna dict en lugar de lanzar excepción.
        """
        if not isinstance(result, dict):
            return None

        success = result.get("success")

        if success is False:
            return (
                result.get("error")
                or result.get("mensaje")
                or "El stage reportó success=False"
            )

        if success is None and result.get("error"):
            return str(result.get("error"))

        return None

    @staticmethod
    def _build_error_details(
        stage_name: str,
        stage_index: int,
        total_stages: int,
        module_path: str,
        function_name: str,
        error: Exception,
        traceback_text: str
    ) -> Dict[str, Any]:
        tb_lines = [line for line in traceback_text.splitlines() if line.strip()]
        return {
            "stage_name": stage_name,
            "stage_index": stage_index + 1,
            "total_stages": total_stages,
            "module_path": module_path,
            "function_name": function_name,
            "exception_type": type(error).__name__,
            "error_message": str(error),
            "traceback_excerpt": tb_lines[:8]
        }
    
    def load_pipeline(self) -> bool:
        """
        Carga la configuración del pipeline desde YAML
        
        Returns:
            True si se cargó correctamente, False en caso contrario
        """
        try:
            with open(self.yaml_path, 'r', encoding='utf-8') as f:
                self.pipeline_config = yaml.safe_load(f)
            
            self._log("INFO", f"✓ Pipeline cargado: {self.pipeline_config['name']}")
            self._log("INFO", f"  • Versión: {self.pipeline_config['version']}")
            self._log("INFO", f"  • Stages: {len(self.pipeline_config['stages'])}")
            
            return True
            
        except Exception as e:
            self._log("ERROR", f"Error al cargar pipeline YAML: {e}")
            return False
    
    def validate_structure(self) -> bool:
        """
        Valida que exista el archivo de entrada y tenga la pestaña 'Practicantes'
        
        Returns:
            True si la estructura es válida
        """
        # Validar que el archivo existe
        if not self.archivo.exists():
            self._log("ERROR", f"No se encontró el archivo: {self.archivo.name}")
            return False
        
        self._log("INFO", f"✓ Archivo encontrado: {self.archivo.name}")
        
        # Warning suave: solo alertar si el filename no trae fecha DD.MM.YYYY
        if not has_document_date_fragment(self.archivo):
            self._log("WARNING", "⚠️  El nombre del archivo no contiene una fecha con formato DD.MM.YYYY")
            self._log("WARNING", f"   Archivo actual: {self.archivo.name}")
        
        # Validar que tenga la pestaña "Practicantes"
        try:
            import openpyxl
            wb = openpyxl.load_workbook(self.archivo, read_only=True)
            
            if "Practicantes" not in wb.sheetnames:
                self._log("ERROR", f"El archivo no contiene la pestaña 'Practicantes'")
                self._log("ERROR", f"Pestañas disponibles: {wb.sheetnames}")
                wb.close()
                return False
            
            wb.close()
            self._log("INFO", f"✓ Pestaña 'Practicantes' encontrada")
            
        except Exception as e:
            self._log("ERROR", f"Error al validar archivo Excel: {e}")
            return False
        
        self._log("INFO", f"✓ Estructura validada correctamente")
        
        return True
    
    def execute_stage(self, stage_config: Dict, stage_index: int, total_stages: int) -> bool:
        """
        Ejecuta un stage individual del pipeline
        
        Args:
            stage_config: Configuración del stage desde YAML
            stage_index: Índice del stage (0-based)
            total_stages: Total de stages en el pipeline
            
        Returns:
            True si el stage se ejecutó correctamente
        """
        stage_name = stage_config['name']
        stage_desc = stage_config.get('description', '')
        
        # Emitir señal de inicio
        self.stage_started.emit(stage_name, stage_desc)
        
        self._log("INFO", "")
        self._log("INFO", "=" * 70)
        self._log("INFO", f"STAGE {stage_index + 1}/{total_stages}: {stage_name}")
        if stage_desc:
            self._log("INFO", f"Descripción: {stage_desc}")
        self._log("INFO", "=" * 70)
        
        tiempo_inicio = time.time()
        module_path = stage_config.get('module', '')
        function_name = stage_config.get('function', '')
        self.last_stage_error = None
        
        try:
            # Importar módulo
            self._log("INFO", f"→ Importando: {module_path}")
            module = import_module(module_path)
            
            if not hasattr(module, function_name):
                raise AttributeError(
                    f"Módulo '{module_path}' no tiene función '{function_name}'"
                )
            
            func = getattr(module, function_name)
            self._log("INFO", f"✓ Función encontrada: {function_name}()")
            
            # Preparar parámetros según el stage
            params = self._prepare_stage_params(stage_config, stage_index)
            
            self._log("INFO", f"→ Ejecutando stage...")
            
            # Calcular progreso base
            progress_base = int((stage_index / total_stages) * 100)
            self._progress(progress_base, f"Ejecutando: {stage_name}")
            
            # Ejecutar función
            result = func(**params)

            stage_error = self._extract_stage_failure(result)
            if stage_error:
                raise RuntimeError(stage_error)
            
            # Guardar resultado para stages posteriores
            self.stages_results[stage_name] = result
            
            # Validar outputs si están definidos
            if 'outputs' in stage_config:
                self._validate_outputs(stage_config['outputs'])
            
            duracion = time.time() - tiempo_inicio
            
            self._log("INFO", f"✓ Stage completado exitosamente")
            self._log("INFO", f"  ⏱️  Duración: {duracion:.2f}s")
            self._log("INFO", "-" * 70)
            
            # Emitir señal de completado
            self.stage_completed.emit(stage_name, True, duracion)
            
            # Actualizar progreso
            progress_end = int(((stage_index + 1) / total_stages) * 100)
            self._progress(progress_end, f"✓ {stage_name}")
            
            return True
            
        except Exception as e:
            duracion = time.time() - tiempo_inicio
            tb_text = traceback.format_exc()
            self.last_stage_error = self._build_error_details(
                stage_name=stage_name,
                stage_index=stage_index,
                total_stages=total_stages,
                module_path=module_path,
                function_name=function_name,
                error=e,
                traceback_text=tb_text
            )

            self._log(
                "ERROR",
                f"Error en stage '{stage_name}' ({module_path}.{function_name}): {str(e)}"
            )

            tb_lines = tb_text.split('\n')
            for line in tb_lines[:10]:
                if line.strip():
                    self._log("DEBUG", f"  {line}")
            
            self.stage_completed.emit(stage_name, False, duracion)
            
            return False
    
    def _prepare_stage_params(self, stage_config: Dict, stage_index: int) -> Dict[str, Any]:
        """
        Prepara parámetros para cada stage según su posición en el pipeline
        
        Args:
            stage_config: Configuración del stage
            stage_index: Índice del stage
            
        Returns:
            Diccionario con parámetros para la función
        """
        if stage_index == 0:  # Stage 1: Bronze → Silver
            # step1_controlpracticantes.procesar_sin_gui(ruta_archivo, carpeta_salida)
            carpeta_silver = self.output_dir / "silver"
            
            return {
                'ruta_archivo': self.archivo,
                'carpeta_salida': carpeta_silver
            }
        
        elif stage_index == 1:  # Stage 2: Silver → Gold
            # step2_controlpracticantes.procesar_sin_gui(ruta_silver, carpeta_gold)
            ruta_silver = self.output_dir / "silver" / "control_practicantes_silver.parquet"
            carpeta_gold = self.output_dir / "gold"
            
            return {
                'ruta_silver': ruta_silver,
                'carpeta_gold': carpeta_gold
            }
        
        else:
            return {}
    
    def _validate_outputs(self, outputs_config: list):
        """
        Valida que los outputs esperados existan
        
        Args:
            outputs_config: Lista de outputs definidos en YAML
        """
        for output in outputs_config:
            path_template = output['path_template']
            required = output.get('required', True)
            
            # Reemplazar variables
            path_str = path_template.replace('${output_dir}', str(self.output_dir))
            output_path = Path(path_str)
            
            if required and not output_path.exists():
                raise FileNotFoundError(f"Output requerido no generado: {output_path}")
            
            if output_path.exists():
                size_mb = output_path.stat().st_size / (1024 * 1024)
                self._log("INFO", f"  ✓ Output generado: {output_path.name} ({size_mb:.2f} MB)")
    
    def execute(self) -> Dict:
        """
        Ejecuta el pipeline completo
        
        Returns:
            Diccionario con resultado de la ejecución
        """
        tiempo_inicio_total = time.time()
        
        # Cargar pipeline
        if not self.load_pipeline():
            return {
                'success': False,
                'error': 'No se pudo cargar el pipeline YAML'
            }
        
        # Validar estructura
        if not self.validate_structure():
            return {
                'success': False,
                'error': 'Validación de archivo fallida'
            }
        
        # Ejecutar stages
        stages = self.pipeline_config['stages']
        total_stages = len(stages)
        
        self._log("INFO", "")
        self._log("INFO", "=" * 70)
        self._log("INFO", f"INICIANDO PIPELINE: {self.pipeline_config['name']}")
        self._log("INFO", f"Total de stages: {total_stages}")
        self._log("INFO", "=" * 70)
        
        failed_stages = []

        for idx, stage_config in enumerate(stages):
            success = self.execute_stage(stage_config, idx, total_stages)
            
            if not success:
                if self.last_stage_error:
                    failed_stages.append(self.last_stage_error.copy())

                if self.pipeline_config['config'].get('stop_on_error', True):
                    self._log("ERROR", "Pipeline detenido por error en stage")
                    error_details = self.last_stage_error or {}
                    error_reason = error_details.get('error_message', 'Error desconocido')
                    return {
                        'success': False,
                        'error': f"Stage '{stage_config['name']}' falló: {error_reason}",
                        'error_details': error_details,
                        'failed_stages': failed_stages,
                        'completed_stages': idx,
                        'total_stages': total_stages,
                        'duracion_total': time.time() - tiempo_inicio_total
                    }

        if failed_stages:
            self._log("ERROR", "Pipeline finalizado con stages fallidos")
            return {
                'success': False,
                'error': f"Pipeline finalizado con {len(failed_stages)} stage(s) fallido(s)",
                'error_details': failed_stages[0],
                'failed_stages': failed_stages,
                'completed_stages': total_stages - len(failed_stages),
                'total_stages': total_stages,
                'duracion_total': time.time() - tiempo_inicio_total,
                'stages_results': self.stages_results
            }
        
        # Pipeline completado
        duracion_total = time.time() - tiempo_inicio_total
        
        self._log("INFO", "")
        self._log("INFO", "=" * 70)
        self._log("INFO", "PIPELINE COMPLETADO EXITOSAMENTE")
        self._log("INFO", "=" * 70)
        self._log("INFO", f"⏱️  Tiempo total: {duracion_total:.2f}s")
        self._log("INFO", "=" * 70)
        
        self._progress(100, "✓ Pipeline completado")
        
        return {
            'success': True,
            'completed_stages': total_stages,
            'total_stages': total_stages,
            'duracion_total': duracion_total,
            'stages_results': self.stages_results
        }


def get_resource_path(relative_path: str) -> Path:
    """
    Obtiene ruta absoluta de recursos
    Compatible con PyInstaller
    """
    try:
        base_path = Path(sys._MEIPASS)
    except AttributeError:
        base_path = Path(__file__).parent.parent
    
    return base_path / relative_path
