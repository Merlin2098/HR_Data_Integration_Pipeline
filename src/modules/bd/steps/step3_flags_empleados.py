"""
Script: step3_flags_empleados.py
Descripción: Generación de Flags en Capa Gold - Empleados
             Aplica reglas de negocio para edad y contratos sobre parquet Gold
             Toda la lógica de cálculo se ejecuta en DuckDB mediante query SQL externa
             
Arquitectura:
- Input: Gold Empleados (bd_empleados_gold.parquet)
- Output: Gold Empleados con Flags

Salida:
- Archivos actuales sin timestamp en gold/
- Copias históricas con timestamp en gold/historico/

Flags generadas:
- tiempo_servicio_texto (String)
- cumple_65_esteaño (Boolean)
- cumple_65_proximoaño (Boolean)
- cumple_70_esteaño (Boolean)
- cumple_70_proximoaño (Boolean)
- alerta_contrato_obra (Boolean)
- alerta_contrato_incremento (Boolean)

Autor: Richi
Fecha: 06.01.2025
Última modificación: 14.01.2025 - Corrección patrón dual /actual + /historico
"""

import polars as pl
import duckdb
from pathlib import Path
from datetime import datetime
import time
from tkinter import Tk, filedialog
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from src.utils.gold_export import maybe_write_excel


def seleccionar_archivo(titulo: str) -> Path | None:
    """Abre un explorador de archivos para seleccionar un archivo."""
    root = Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    
    archivo = filedialog.askopenfilename(
        title=titulo,
        filetypes=[
            ("Parquet files", "*.parquet"),
            ("All files", "*.*")
        ]
    )
    
    root.destroy()
    
    return Path(archivo) if archivo else None


def buscar_carpeta_queries() -> Path | None:
    """
    Busca la carpeta 'queries' en el directorio actual y hasta 3 niveles arriba.
    
    Returns:
        Path de la carpeta queries o None si no se encuentra
    """
    carpeta_actual = Path.cwd()
    
    # Buscar en el directorio actual y hasta 3 niveles arriba
    for _ in range(4):
        posible_queries = carpeta_actual / "queries"
        if posible_queries.exists() and posible_queries.is_dir():
            return posible_queries
        carpeta_actual = carpeta_actual.parent
    
    # También buscar en el directorio del script
    script_dir = Path(__file__).parent
    for _ in range(4):
        posible_queries = script_dir / "queries"
        if posible_queries.exists() and posible_queries.is_dir():
            return posible_queries
        script_dir = script_dir.parent
    
    return None


def cargar_query_sql() -> str:
    """
    Busca la carpeta 'queries' y carga el archivo queries_flags_gold.sql
    
    Returns:
        String con el contenido del archivo SQL
    """
    print("\n📄 Buscando carpeta de queries...")
    
    # Buscar carpeta de queries
    carpeta_queries = buscar_carpeta_queries()
    
    if carpeta_queries is None:
        raise FileNotFoundError(
            "No se encontró la carpeta 'queries' en el proyecto.\n"
            "Asegúrate de que exista la carpeta 'queries' en la raíz del proyecto."
        )
    
    print(f"  ✓ Carpeta encontrada: {carpeta_queries}")
    
    # Buscar el archivo queries_flags_gold.sql
    ruta_query = carpeta_queries / "queries_flags_gold.sql"
    
    if not ruta_query.exists():
        # Listar archivos SQL disponibles
        queries_disponibles = list(carpeta_queries.glob("*.sql"))
        mensaje = f"No se encontró 'queries_flags_gold.sql' en {carpeta_queries}\n"
        
        if queries_disponibles:
            mensaje += "\nArchivos SQL disponibles:\n"
            for query_file in queries_disponibles:
                mensaje += f"  • {query_file.name}\n"
            mensaje += "\nAsegúrate de que el archivo se llame 'queries_flags_gold.sql'"
        else:
            mensaje += "No hay archivos SQL en la carpeta queries"
        
        raise FileNotFoundError(mensaje)
    
    print(f"  ✓ Query encontrada: {ruta_query.name}")
    
    with open(ruta_query, 'r', encoding='utf-8') as f:
        query = f.read()
    
    print(f"  ✓ Query cargada ({len(query)} caracteres)")
    
    return query


def validar_columnas_requeridas(df: pl.DataFrame) -> bool:
    """
    Valida que el DataFrame tenga las columnas requeridas para aplicar flags.
    
    Args:
        df: DataFrame de Polars
    
    Returns:
        True si todas las columnas requeridas existen
    """
    columnas_requeridas = [
        "NUMERO DE DOC",
        "FECHA DE NAC.",
        "FECH_INGR.",
        "Fecha de Termino",
        "Modalidad de Contrato"
    ]
    
    columnas_presentes = df.columns
    columnas_faltantes = [col for col in columnas_requeridas if col not in columnas_presentes]
    
    if columnas_faltantes:
        print(f"\n  ❌ COLUMNAS FALTANTES:")
        for col in columnas_faltantes:
            print(f"     • {col}")
        return False
    
    print(f"  ✓ Todas las columnas requeridas están presentes")
    return True


def aplicar_flags_duckdb(df: pl.DataFrame, query: str) -> pl.DataFrame:
    """
    Ejecuta las queries SQL en DuckDB para calcular todas las flags.
    
    Args:
        df: DataFrame de Polars con datos originales
        query: String con la query SQL completa
    
    Returns:
        DataFrame con las flags aplicadas
    """
    print(f"\n  📊 Ejecutando queries SQL en DuckDB...")
    print("  " + "-" * 76)
    
    # Conectar a DuckDB
    con = duckdb.connect(":memory:")
    
    # Registrar el DataFrame de Polars en DuckDB
    con.register("empleados", df)
    
    # Ejecutar query y obtener resultado
    df_con_flags = con.execute(query).pl()
    
    # Contar flags aplicadas (solo booleanas)
    flags_cols = [col for col in df_con_flags.columns if col not in df.columns]
    
    print(f"  ✓ Query ejecutada exitosamente")
    print(f"  ✓ Columnas de flags generadas: {len(flags_cols)}")
    
    for flag_col in flags_cols:
        # Verificar si es booleana antes de hacer sum
        if df_con_flags[flag_col].dtype == pl.Boolean:
            count = df_con_flags[flag_col].sum()
            total = len(df_con_flags)
            porcentaje = (count / total * 100) if total > 0 else 0
            print(f"     • {flag_col}: {count} empleados ({porcentaje:.2f}%)")
        else:
            # Para columnas no booleanas (como tiempo_servicio_texto)
            print(f"     • {flag_col}: columna de texto generada")
    
    # Cerrar conexión
    con.close()
    
    print("  " + "-" * 76)
    
    return df_con_flags


def generar_resumen_flags(df: pl.DataFrame, columnas_originales: list) -> dict:
    """
    Genera un resumen estadístico de las flags aplicadas.
    
    Args:
        df: DataFrame con flags aplicadas
        columnas_originales: Lista de columnas originales (antes de flags)
    
    Returns:
        Diccionario con estadísticas
    """
    stats = {
        "total_registros": len(df),
        "columnas_originales": len(columnas_originales),
        "columnas_totales": len(df.columns),
        "flags_generadas": {},
        "modalidades": {}
    }
    
    # Identificar columnas de flags (nuevas columnas)
    flags_cols = [col for col in df.columns if col not in columnas_originales]
    
    # Contar empleados por cada flag booleana
    for flag_col in flags_cols:
        if df[flag_col].dtype == pl.Boolean:
            count = df[flag_col].sum()
            stats["flags_generadas"][flag_col] = {
                "count": count,
                "porcentaje": round(count / len(df) * 100, 2)
            }
    
    # Distribución por modalidad
    if "Modalidad de Contrato" in df.columns:
        modalidades = df.group_by("Modalidad de Contrato").agg(
            pl.len().alias("cantidad")
        ).sort("cantidad", descending=True)
        
        for row in modalidades.iter_rows(named=True):
            modalidad = row["Modalidad de Contrato"]
            cantidad = row["cantidad"]
            stats["modalidades"][modalidad] = {
                "count": cantidad,
                "porcentaje": round(cantidad / len(df) * 100, 2)
            }
    
    return stats


def guardar_resultados(
    df: pl.DataFrame,
    carpeta_gold: Path,
    export_excel: bool = False,
) -> tuple[Path, Path | None, Path, Path | None]:
    """
    Guarda el resultado con versionamiento:
    - Archivos actuales sin timestamp en gold/
    - Copia con timestamp en gold/historico/
    
    Args:
        df: DataFrame con flags aplicadas
        carpeta_gold: Path de la carpeta gold
    
    Returns:
        tuple: (actual_parquet, actual_excel, historico_parquet, historico_excel)
    """
    print(f"\n[3/3] Guardando resultados...")
    
    # Crear carpeta historico/ si no existe
    carpeta_historico = carpeta_gold / "historico"
    carpeta_historico.mkdir(exist_ok=True)
    
    # Timestamp para archivo histórico
    timestamp = datetime.now().strftime("%d.%m.%Y_%H.%M.%S")
    
    print(f"  📁 Carpeta Gold: {carpeta_gold}")
    print(f"  📁 Carpeta Histórico: {carpeta_historico}")
    
    # === ARCHIVOS ACTUALES (sin timestamp) ===
    nombre_actual = "bd_empleados_flags_gold"
    ruta_parquet_actual = carpeta_gold / f"{nombre_actual}.parquet"
    ruta_excel_actual = carpeta_gold / f"{nombre_actual}.xlsx"
    
    print(f"\n  📄 Archivos actuales (se sobreescriben):")
    print(f"    - Guardando parquet...", end='', flush=True)
    df.write_parquet(ruta_parquet_actual, compression="snappy")
    tamanio_mb = ruta_parquet_actual.stat().st_size / (1024 * 1024)
    print(f" ✓ ({tamanio_mb:.2f} MB)")
    
    # Guardar Excel usando openpyxl directamente (más eficiente)
    wb = Workbook()
    ws = wb.active
    ws.title = "Empleados_Flags"
    
    # Formato de encabezado
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Escribir encabezados
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Escribir datos fila por fila
    for row_idx, row_data in enumerate(df.iter_rows(), 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Ajustar ancho de columnas
    for col_idx, col_name in enumerate(df.columns, 1):
        max_length = len(str(col_name))
        sample_values = df[col_name].head(100)
        for val in sample_values:
            val_len = len(str(val))
            if val_len > max_length:
                max_length = val_len
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = adjusted_width
    
    ruta_excel_actual = maybe_write_excel(
        ruta_excel_actual,
        export_excel,
        lambda path: wb.save(path),
    )
    if ruta_excel_actual is not None:
        tamanio_mb = ruta_excel_actual.stat().st_size / (1024 * 1024)
        print(f"    - Guardando Excel...", end='', flush=True)
        print(f" ✓ ({tamanio_mb:.2f} MB)")
    else:
        print("    - Excel actual omitido (exportación opcional desactivada)")
    
    # === ARCHIVOS HISTÓRICOS (con timestamp) ===
    nombre_historico = f"bd_empleados_flags_gold_{timestamp}"
    ruta_parquet_historico = carpeta_historico / f"{nombre_historico}.parquet"
    ruta_excel_historico = carpeta_historico / f"{nombre_historico}.xlsx"
    
    print(f"\n  📦 Archivos históricos (con timestamp):")
    print(f"    - Guardando parquet...", end='', flush=True)
    df.write_parquet(ruta_parquet_historico, compression="snappy")
    print(f" ✓")
    
    ruta_excel_historico = maybe_write_excel(
        ruta_excel_historico,
        export_excel,
        lambda path: wb.save(path),
    )
    if ruta_excel_historico is not None:
        print(f"    - Guardando Excel...", end='', flush=True)
        print(f" ✓")
    else:
        print("    - Excel histórico omitido (exportación opcional desactivada)")
    
    return ruta_parquet_actual, ruta_excel_actual, ruta_parquet_historico, ruta_excel_historico


def main():
    """Función principal del script"""
    print("=" * 80)
    print(" GENERACIÓN DE FLAGS - EMPLEADOS ".center(80, "="))
    print("=" * 80)
    
    # 1. Seleccionar archivo Parquet Gold
    print("\n[PASO 1] Selecciona el archivo Parquet Gold de Empleados...")
    input_file = seleccionar_archivo("Seleccionar Parquet Gold - Empleados")
    
    if not input_file:
        print("\n❌ No se seleccionó ningún archivo. Proceso cancelado.")
        return
    
    if not input_file.exists():
        print(f"\n❌ El archivo no existe: {input_file}")
        return
    
    # Iniciar temporizador DESPUÉS de seleccionar archivo
    start_time = time.time()
    
    print(f"✓ Archivo seleccionado: {input_file.name}")
    
    # 2. Cargar query SQL
    print(f"\n[1/3] Cargando query SQL...")
    try:
        query = cargar_query_sql()
    except Exception as e:
        print(f"\n❌ Error al cargar query SQL: {e}")
        return
    
    # 3. Cargar Parquet Gold
    print(f"\n[2/3] Procesando datos...")
    print(f"  📂 Cargando parquet: {input_file.name}")
    try:
        df = pl.read_parquet(input_file)
        print(f"  ✓ Datos cargados: {df.shape[0]:,} filas × {df.shape[1]} columnas")
        columnas_originales = df.columns.copy()
    except Exception as e:
        print(f"\n❌ Error al cargar Parquet: {e}")
        return
    
    # 4. Validar columnas requeridas
    print(f"\n  🔍 Validando columnas requeridas...")
    if not validar_columnas_requeridas(df):
        print(f"\n❌ El archivo no contiene las columnas requeridas. Proceso cancelado.")
        return
    
    # 5. Aplicar flags mediante DuckDB
    try:
        df_con_flags = aplicar_flags_duckdb(df, query)
        print(f"\n  ✓ Dataset resultante: {df_con_flags.shape[0]:,} filas × {df_con_flags.shape[1]} columnas")
    except Exception as e:
        print(f"\n❌ Error al aplicar flags: {e}")
        import traceback
        traceback.print_exc()
        return
    
    # 6. Generar resumen de flags
    print(f"\n  📈 Generando resumen de flags...")
    stats = generar_resumen_flags(df_con_flags, columnas_originales)
    
    # 7. Guardar archivos
    try:
        carpeta_gold = input_file.parent
        ruta_p_act, ruta_e_act, ruta_p_hist, ruta_e_hist = guardar_resultados(df_con_flags, carpeta_gold)
    except Exception as e:
        print(f"\n❌ Error al guardar archivos: {e}")
        import traceback
        traceback.print_exc()
        return
    
    # Calcular tiempo de ejecución
    elapsed_time = time.time() - start_time
    
    # 8. Resumen final
    print("\n" + "=" * 80)
    print(" RESUMEN ".center(80, "="))
    print("=" * 80)
    
    print(f"\n✓ Proceso completado exitosamente")
    
    print(f"\n📊 Estadísticas Generales:")
    print(f"  - Total registros procesados: {stats['total_registros']:,}")
    print(f"  - Columnas originales: {stats['columnas_originales']}")
    print(f"  - Columnas totales: {stats['columnas_totales']}")
    print(f"  - Flags generadas: {len(stats['flags_generadas'])}")
    print(f"  - Tiempo de ejecución: {elapsed_time:.2f}s")
    
    print(f"\n🏷️  Flags Aplicadas:")
    for flag_name, flag_data in stats["flags_generadas"].items():
        print(f"  • {flag_name}: {flag_data['count']:,} empleados ({flag_data['porcentaje']}%)")
    
    if stats["modalidades"]:
        print(f"\n🏢 Distribución por Modalidad de Contrato:")
        for modalidad, modalidad_data in stats["modalidades"].items():
            print(f"  • {modalidad}: {modalidad_data['count']:,} ({modalidad_data['porcentaje']}%)")
    
    print(f"\n📁 Archivos generados:")
    print(f"\n  Actuales (para Power BI):")
    print(f"    - {ruta_p_act.name}")
    if ruta_e_act is not None:
        print(f"    - {ruta_e_act.name}")
    
    print(f"\n  Históricos (con timestamp):")
    print(f"    - {ruta_p_hist.name}")
    if ruta_e_hist is not None:
        print(f"    - {ruta_e_hist.name}")
    
    print(f"\n📂 Ubicación: {carpeta_gold}")
    
    print("\n💡 Notas:")
    print("  - Archivos actuales: se sobreescriben (rutas estables para Power BI)")
    print("  - Archivos históricos: se archivan con timestamp en gold/historico/")
    
    print("\n" + "=" * 80)


if __name__ == "__main__":
    main()
