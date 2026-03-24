"""
Script: step1_capasilver.py
Descripción: Procesa archivo Excel Bronze → Parquet Silver
             - Encabezados en fila 10
             - Datos desde fila 11
             - Identifica última fila mediante columna NUMERO DE DOC

Arquitectura:
- Bronze: Excel con datos de BD
- Silver: Parquet limpio y estandarizado (SE SOBRESCRIBE EN CADA EJECUCIÓN)

Salida: Archivos sin timestamp en carpeta silver/
    - bd_silver.parquet

Autor: Richi
Fecha: 06.01.2025
"""

import polars as pl
import openpyxl
from pathlib import Path
from datetime import datetime
import time
from tkinter import Tk, filedialog
import re

from src.utils.bd_document_date import append_bd_document_date_column, extract_bd_document_date


def seleccionar_archivo_excel() -> Path | None:
    """Abre diálogo para seleccionar archivo Excel"""
    root = Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    
    archivo = filedialog.askopenfilename(
        title="Seleccionar archivo Excel Bronze - BD",
        filetypes=[("Excel files", "*.xlsx *.xlsm *.xls"), ("All files", "*.*")]
    )
    
    root.destroy()
    
    return Path(archivo) if archivo else None


def extraer_datos_excel(file_path: Path, header_row: int = 10, data_start_row: int = 11) -> tuple[list[str], list[list]]:
    """
    Extrae encabezados y datos del archivo Excel usando openpyxl.
    Convierte strings en formato DD/MM/YYYY a datetime para estandarización.
    OPTIMIZADO: Usa iter_rows para mejor rendimiento
    
    Args:
        file_path: Ruta del archivo Excel
        header_row: Número de fila donde están los encabezados (1-indexed)
        data_start_row: Número de fila donde inician los datos (1-indexed)
    
    Returns:
        Tupla con (lista de encabezados, lista de filas de datos)
    """
    print(f"\n📂 Cargando archivo: {file_path.name}")
    
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    ws = wb.active
    
    print(f"📄 Hoja activa: {ws.title}")
    
    # Extraer encabezados de la fila especificada
    headers = []
    for cell in ws[header_row]:
        if cell.value:
            headers.append(str(cell.value))
        else:
            break
    
    print(f"📊 Encabezados detectados: {len(headers)}")
    
    # Buscar la columna NUMERO DE DOC
    numero_doc_col_idx = None
    for idx, header in enumerate(headers, 1):
        if "NUMERO" in header.upper() and "DOC" in header.upper():
            numero_doc_col_idx = idx
            print(f"🔍 Columna 'NUMERO DE DOC' encontrada en posición {idx}: {header}")
            break
    
    if not numero_doc_col_idx:
        raise ValueError("No se encontró la columna 'NUMERO DE DOC' en los encabezados")
    
    # Patrón para detectar fechas en formato DD/MM/YYYY o D/M/YYYY
    date_pattern = re.compile(r'^\d{1,2}/\d{1,2}/\d{4}$')
    
    # Contador de conversiones
    conversiones_fecha = 0
    
    # Extraer datos usando iter_rows para mejor rendimiento
    print(f"\n📊 Extrayendo datos desde fila {data_start_row}...")
    data_rows = []
    
    for row in ws.iter_rows(min_row=data_start_row, max_col=len(headers), values_only=True):
        # Verificar si la columna NUMERO DE DOC tiene valor
        if row[numero_doc_col_idx - 1] is None or str(row[numero_doc_col_idx - 1]).strip() == "":
            break
        
        row_data = []
        for cell_value in row:
            # Si el valor es un string que parece una fecha en formato DD/MM/YYYY
            if isinstance(cell_value, str):
                cell_value = cell_value.strip()
                
                if date_pattern.match(cell_value):
                    try:
                        # Convertir DD/MM/YYYY a datetime
                        cell_value = datetime.strptime(cell_value, "%d/%m/%Y")
                        conversiones_fecha += 1
                    except ValueError:
                        pass
            
            row_data.append(cell_value)
        
        data_rows.append(row_data)
    
    wb.close()
    
    print(f"✓ Filas de datos extraídas: {len(data_rows):,}")
    if conversiones_fecha > 0:
        print(f"✓ Fechas convertidas de DD/MM/YYYY: {conversiones_fecha}")
    
    return headers, data_rows


def crear_dataframe_polars(headers: list[str], data_rows: list[list]) -> pl.DataFrame:
    """
    Crea un DataFrame de Polars a partir de encabezados y datos.
    Convierte datetime a string con formato consistente YYYY-MM-DD HH:MM:SS.
    
    Args:
        headers: Lista de nombres de columnas
        data_rows: Lista de filas de datos
    
    Returns:
        DataFrame de Polars
    """
    print(f"\n🔄 Creando DataFrame de Polars...")
    
    # Procesar data_rows para convertir datetime a string con formato consistente
    processed_rows = []
    for row in data_rows:
        processed_row = []
        for value in row:
            # Convertir datetime a string con formato consistente YYYY-MM-DD HH:MM:SS
            if isinstance(value, datetime):
                processed_row.append(value.strftime("%Y-%m-%d %H:%M:%S"))
            else:
                processed_row.append(value)
        processed_rows.append(processed_row)
    
    # Crear diccionario para DataFrame
    data_dict = {header: [row[i] if i < len(row) else None for row in processed_rows] 
                 for i, header in enumerate(headers)}
    
    # Convertir todas las columnas a string para evitar conflictos de tipo
    for key in data_dict:
        data_dict[key] = [str(v) if v is not None else None for v in data_dict[key]]
    
    # Reemplazar 'None' y 'nan' strings por valores nulos reales
    for key in data_dict:
        data_dict[key] = [None if v in ['None', 'nan', 'NaT'] else v for v in data_dict[key]]
    
    # Crear DataFrame con strict=False
    df = pl.DataFrame(data_dict, strict=False)
    
    print(f"✓ DataFrame creado: {df.height:,} filas × {df.width} columnas")
    
    return df


def guardar_resultados(df: pl.DataFrame, carpeta_trabajo: Path):
    """
    Guarda el DataFrame Silver como parquet en carpeta silver/
    Sin timestamp - se sobreescribe en cada ejecución
    
    Args:
        df: DataFrame a guardar
        carpeta_trabajo: Path de la carpeta de trabajo
        
    Returns:
        Path: ruta del parquet generado
    """
    # Crear carpeta silver si no existe
    carpeta_silver = carpeta_trabajo / "silver"
    carpeta_silver.mkdir(exist_ok=True)
    
    # Nombres fijos sin timestamp
    nombre_base = "bd_silver"
    ruta_parquet = carpeta_silver / f"{nombre_base}.parquet"
    
    print(f"\n[2/2] Guardando resultados en capa Silver...")
    print(f"  📁 Carpeta: {carpeta_silver}")
    
    # Guardar Parquet
    print(f"  - Guardando parquet...", end='', flush=True)
    df.write_parquet(ruta_parquet, compression="snappy")
    print(f" ✓")
    print(f"    Ubicación: {ruta_parquet.name}")
    
    return ruta_parquet


def main():
    """Función principal del script"""
    print("=" * 80)
    print(" PROCESADOR BD - BRONZE → SILVER ".center(80, "="))
    print("=" * 80)
    
    # 1. Seleccionar archivo
    print("\n[PASO 1] Selecciona el archivo Excel Bronze (BD)...")
    archivo_bronze = seleccionar_archivo_excel()
    
    if not archivo_bronze:
        print("❌ No se seleccionó ningún archivo. Proceso cancelado.")
        return
    
    if not archivo_bronze.exists():
        print(f"❌ El archivo no existe: {archivo_bronze}")
        return
    
    # Iniciar cronómetro después de la selección
    tiempo_inicio = time.time()
    
    print(f"✓ Archivo seleccionado: {archivo_bronze.name}")
    carpeta_trabajo = archivo_bronze.parent
    
    # 2. Procesar datos
    print("\n" + "=" * 80)
    print(" PROCESAMIENTO ".center(80, "="))
    print("=" * 80)
    print(f"\n[1/2] Extrayendo datos del Excel...")
    
    try:
        # 2.1 Extraer datos
        headers, data_rows = extraer_datos_excel(
            archivo_bronze,
            header_row=10,
            data_start_row=11
        )
        
        # 2.2 Crear DataFrame
        df = crear_dataframe_polars(headers, data_rows)

        # 2.3 Agregar metadata del documento desde el filename
        fecha_documento = extract_bd_document_date(archivo_bronze)
        df = append_bd_document_date_column(df, archivo_bronze)
        print(f"✓ FECHA_DOCUMENTO detectada desde filename: {fecha_documento}")
        
        # 3. Guardar resultados
        ruta_parquet = guardar_resultados(df, carpeta_trabajo)
        
        # Calcular tiempo total
        tiempo_total = time.time() - tiempo_inicio
        
        # 4. Resumen final
        print("\n" + "=" * 80)
        print(" RESUMEN ".center(80, "="))
        print("=" * 80)
        
        print(f"\n✓ Procesamiento completado exitosamente")
        print(f"\n📊 Estadísticas:")
        print(f"  - Total de registros: {df.height:,}")
        print(f"  - Total de columnas: {df.width}")
        
        print(f"\n📁 Archivos generados en carpeta silver/:")
        print(f"  - Parquet: {ruta_parquet.name}")
        
        print(f"\n⏱️  Tiempo de ejecución: {tiempo_total:.2f}s")
        
        print("\n💡 Los archivos se sobreescriben en cada ejecución (sin historial)")
        
        print("\n" + "=" * 80)
        
    except Exception as e:
        print(f"\n❌ Error durante el procesamiento: {str(e)}")
        import traceback
        traceback.print_exc()
        return


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"\n❌ Error fatal: {str(e)}")
        import traceback
        traceback.print_exc()
