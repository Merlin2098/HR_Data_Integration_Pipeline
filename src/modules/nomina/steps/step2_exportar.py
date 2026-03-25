"""
Script de transformación Silver → Gold para reportes de planilla
Guarda salida Gold únicamente en carpeta actual/

REFACTORIZADO para compatibilidad con worker UI y estructura simplificada
"""

import polars as pl
from pathlib import Path
from datetime import datetime
import traceback
import os
import time

from src.utils.structured_config import (
    list_structured_files,
    load_structured_data,
    resolve_structured_path,
    structured_filetypes,
)
from src.utils.gold_export import maybe_write_excel
from src.utils.month_name import add_month_name_column


def aplicar_transformaciones_gold(df, schema):
    """Aplica transformaciones según el esquema gold"""
    
    # NO normalizar columnas - mantener nombres originales del parquet
    print(f"\n📋 Seleccionando columnas del schema...")
    print("-" * 70)
    
    # Seleccionar solo las columnas que existen en el schema
    columnas_schema = list(schema["schema"].keys())
    columnas_disponibles = [col for col in columnas_schema if col in df.columns]
    columnas_faltantes = [col for col in columnas_schema if col not in df.columns]
    
    for col in columnas_disponibles:
        print(f"✓ {col}")
    
    if columnas_faltantes:
        print(f"\n⚠️  Columnas del schema no encontradas en el parquet:")
        for col in columnas_faltantes:
            print(f"  ✗ {col}")
    
    print("-" * 70)
    print(f"Total: {len(columnas_disponibles)} de {len(columnas_schema)} columnas")
    print()
    
    # Seleccionar columnas
    df = df.select(columnas_disponibles)
    
    # Aplicar tipos de datos según schema
    print("🔄 Aplicando tipos de datos...")
    for col_name in columnas_disponibles:
        col_spec = schema["schema"][col_name]
        col_type = col_spec["type"]
        
        try:
            if col_type == "string":
                df = df.with_columns(pl.col(col_name).cast(pl.Utf8))
            elif col_type == "integer":
                df = df.with_columns(pl.col(col_name).cast(pl.Int64))
            elif col_type == "float":
                df = df.with_columns(pl.col(col_name).cast(pl.Float64))
            elif col_type == "date":
                # Manejar fechas que vienen como datetime string
                df = df.with_columns(
                    pl.col(col_name).str.strptime(pl.Datetime, "%Y-%m-%d %H:%M:%S%.f", strict=False).cast(pl.Date)
                )
        except Exception as e:
            print(f"  ⚠️  Error al convertir columna {col_name}: {e}")
    
    # Aplicar string trimming
    print("✂️  Limpiando espacios en strings...")
    for col in df.columns:
        if df[col].dtype == pl.Utf8:
            df = df.with_columns(pl.col(col).str.strip_chars())
    
    return df


def agregar_nombre_mes(df):
    """Agrega columna con nombre del mes en inglés."""
    
    print("📅 Agregando nombre del mes...")

    df = add_month_name_column(df, default_invalid=None)
    
    print(f"✓ Columna 'NOMBRE_MES' agregada después de 'MES'")
    
    return df


def validar_constraints(df, schema):
    """Valida constraints del schema"""
    errores = []
    warnings = []
    
    print("\n✅ Validando constraints...")
    print("-" * 70)
    
    # Validar primary key
    pk_cols = schema["constraints"]["primary_key"]
    pk_cols_existentes = [col for col in pk_cols if col in df.columns]
    
    if len(pk_cols_existentes) != len(pk_cols):
        faltantes = set(pk_cols) - set(pk_cols_existentes)
        warnings.append(f"⚠️  Columnas de primary key faltantes: {faltantes}")
        pk_cols = pk_cols_existentes
    
    if pk_cols:
        duplicados = df.group_by(pk_cols).agg(pl.len().alias("count")).filter(pl.col("count") > 1)
        
        if len(duplicados) > 0:
            errores.append(f"❌ Se encontraron {len(duplicados)} registros duplicados en primary key: {pk_cols}")
        else:
            print(f"✓ Primary key sin duplicados: {pk_cols}")
    
    # Validar nulls en columnas no nullable
    nulls_encontrados = False
    for col_name, col_spec in schema["schema"].items():
        if col_name not in df.columns:
            continue
            
        if not col_spec.get("nullable", True):
            nulls = df.filter(pl.col(col_name).is_null()).height
            if nulls > 0:
                errores.append(f"❌ Columna '{col_name}' tiene {nulls} valores nulos (no permitidos)")
                nulls_encontrados = True
    
    if not nulls_encontrados:
        print("✓ Validación de nulls correcta")
    
    # Validar valores permitidos
    valores_invalidos_encontrados = False
    for col_name, col_spec in schema["schema"].items():
        if col_name not in df.columns:
            continue
            
        if "allowed_values" in col_spec:
            valores_invalidos = df.filter(
                ~pl.col(col_name).is_in(col_spec["allowed_values"]) & 
                pl.col(col_name).is_not_null()
            )
            if len(valores_invalidos) > 0:
                valores_unicos = valores_invalidos[col_name].unique().to_list()[:5]
                warnings.append(f"⚠️  Columna '{col_name}' tiene {len(valores_invalidos)} valores fuera del rango permitido: {valores_unicos}")
                valores_invalidos_encontrados = True
    
    if not valores_invalidos_encontrados:
        print("✓ Validación de valores permitidos correcta")
    
    # Validar rangos numéricos
    rangos_invalidos_encontrados = False
    for col_name, col_spec in schema["schema"].items():
        if col_name not in df.columns:
            continue
            
        if "min_value" in col_spec:
            fuera_rango = df.filter(
                (pl.col(col_name) < col_spec["min_value"]) & 
                pl.col(col_name).is_not_null()
            ).height
            if fuera_rango > 0:
                warnings.append(f"⚠️  Columna '{col_name}' tiene {fuera_rango} valores menores al mínimo permitido ({col_spec['min_value']})")
                rangos_invalidos_encontrados = True
    
    if not rangos_invalidos_encontrados:
        print("✓ Validación de rangos numéricos correcta")
    
    print("-" * 70)
    
    return errores, warnings


def generar_excel_visualizacion(df, ruta_salida):
    """Genera Excel con formato para visualización"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    print(f"  - Generando Excel con formato...")
    
    # Convertir a pandas para openpyxl
    df_pandas = df.to_pandas()
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Planilla Gold"
    
    # Estilos
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    border_style = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )
    
    # Escribir encabezados
    for col_idx, column in enumerate(df_pandas.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=column)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border_style
    
    # Escribir datos
    for row_idx, row in enumerate(df_pandas.values, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border_style
            
            # Alineación según tipo de dato
            if isinstance(value, (int, float)):
                cell.alignment = Alignment(horizontal='right')
            else:
                cell.alignment = Alignment(horizontal='left')
    
    # Ajustar ancho de columnas
    for column_cells in ws.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        adjusted_width = min(length + 2, 50)
        ws.column_dimensions[column_cells[0].column_letter].width = adjusted_width
    
    # Congelar primera fila
    ws.freeze_panes = 'A2'
    
    # Guardar
    wb.save(ruta_salida)
    print(f"    ✓ Excel generado con formato")


def gestionar_versionamiento_gold(carpeta_base):
    """
    Prepara la estructura de salida para archivos gold.
    ESTRUCTURA SIMPLIFICADA:
    gold/
    ├── actual/        <- Power BI apunta aquí
    
    Args:
        carpeta_base: Carpeta raíz del proyecto (donde está silver/)
        
    Returns:
        Path: carpeta_actual
    """
    carpeta_gold = Path(carpeta_base) / "gold"
    carpeta_actual = carpeta_gold / "actual"      # ← SIN subcarpeta "nomina"
    
    # Crear estructura si no existe
    carpeta_actual.mkdir(parents=True, exist_ok=True)
    
    return carpeta_actual


# ============================================================================
# FUNCIONES PARA COMPATIBILIDAD CON WORKER UI
# ============================================================================

def seleccionar_y_convertir_columnas(df_silver, esquema):
    """
    Función de compatibilidad para el worker UI
    Aplica todas las transformaciones gold en un solo paso
    
    Args:
        df_silver: DataFrame de la capa silver
        esquema: Diccionario con el esquema gold cargado desde YAML
        
    Returns:
        pl.DataFrame: DataFrame transformado a gold
    """
    # Aplicar transformaciones del esquema
    df_gold = aplicar_transformaciones_gold(df_silver, esquema)
    
    # Agregar columna NOMBRE_MES
    df_gold = agregar_nombre_mes(df_gold)
    
    # Validar constraints (solo warnings, no detiene ejecución)
    errores, warnings = validar_constraints(df_gold, esquema)
    
    if errores:
        print("\n⚠️  ERRORES CRÍTICOS ENCONTRADOS:")
        for error in errores:
            print(f"  {error}")
    
    if warnings:
        print("\n⚠️  ADVERTENCIAS:")
        for warning in warnings:
            print(f"  {warning}")
    
    return df_gold


def guardar_resultados(df_gold, carpeta_silver, export_excel: bool = False):
    """
    Función de compatibilidad para el worker UI
    Guarda los archivos gold en actual/
    
    Args:
        df_gold: DataFrame gold procesado
        carpeta_silver: Path de la carpeta silver (se usa para encontrar la base)
        
    Returns:
        dict: Diccionario con las rutas de los archivos generados
    """
    # Subir desde silver/ a carpeta base del proyecto
    carpeta_base = Path(carpeta_silver).parent
    
    # Preparar salida gold
    carpeta_actual = gestionar_versionamiento_gold(carpeta_base)
    
    # Rutas de salida - NOMBRES QUE EL PIPELINE ESPERA
    ruta_parquet_gold = carpeta_actual / "Planilla_Metso_Consolidado.parquet"
    ruta_excel_gold = carpeta_actual / "Planilla_Metso_Consolidado.xlsx"
    
    # Guardar archivos
    print("\n💾 Guardando archivos en capa Gold...")
    print(f"  📁 Carpeta actual: {carpeta_actual}")
    print("-" * 70)
    
    # Guardar parquet gold
    try:
        df_gold.write_parquet(ruta_parquet_gold)
        size_mb = ruta_parquet_gold.stat().st_size / (1024 * 1024) if ruta_parquet_gold.exists() else 0
        print(f"✓ Parquet gold: {ruta_parquet_gold.name} ({size_mb:.2f} MB)")
    except Exception as e:
        print(f"✗ ERROR al guardar parquet: {e}")
        traceback.print_exc()
        raise
    
    # Generar Excel de visualización
    try:
        ruta_excel_gold = maybe_write_excel(
            ruta_excel_gold,
            export_excel,
            lambda path: generar_excel_visualizacion(df_gold, path),
        )
        if ruta_excel_gold is not None:
            size_mb = ruta_excel_gold.stat().st_size / (1024 * 1024) if ruta_excel_gold.exists() else 0
            print(f"✓ Excel gold: {ruta_excel_gold.name} ({size_mb:.2f} MB)")
        else:
            print("ℹ️  Excel gold omitido (exportación opcional desactivada)")
    except Exception as e:
        print(f"⚠️  Error al generar Excel: {e}")
        ruta_excel_gold = None
    
    print("-" * 70)
    
    return {
        'parquet': ruta_parquet_gold,
        'excel': ruta_excel_gold,
        'carpeta_actual': carpeta_actual
    }


def exportar_a_gold(
    ruta_parquet_silver: Path,
    carpeta_trabajo: Path,
    export_excel_gold: bool = False,
) -> dict:
    """
    Procesa Silver a Gold sin interfaz gráfica (modo headless)
    Usado por el pipeline executor
    Genera estructura SIMPLIFICADA: gold/actual/
    
    Args:
        ruta_parquet_silver: Path al parquet Silver de nómina
        carpeta_trabajo: Path a la carpeta de trabajo base
        
    Returns:
        dict con resultados del procesamiento
    """
    
        # ==================== DEBUG ====================
    print(f"🔍 DEBUG - Iniciando exportar_a_gold()")
    print(f"🔍 DEBUG - ruta_parquet_silver: {ruta_parquet_silver}")
    print(f"🔍 DEBUG - carpeta_trabajo: {carpeta_trabajo}")
    print(f"🔍 DEBUG - ¿Silver existe?: {os.path.exists(ruta_parquet_silver)}")
    
    if not os.path.exists(ruta_parquet_silver):
        print(f"🔍 DEBUG - ❌ ARCHIVO NO ENCONTRADO")
        # Listar archivos en la carpeta silver
        silver_dir = os.path.dirname(ruta_parquet_silver)
        print(f"🔍 DEBUG - Contenido de {silver_dir}:")
        if os.path.exists(silver_dir):
            for file in os.listdir(silver_dir):
                print(f"  • {file}")
        else:
            print(f"  • Carpeta {silver_dir} no existe")
    # ==================== FIN DEBUG ====================
    
    # Iniciar temporizador
    tiempo_inicio = time.time()
    
    
    
    print(f"\n🔄 Procesando Silver → Gold (modo headless)...")
    print(f"   Silver: {ruta_parquet_silver.name}")
    print(f"   Carpeta trabajo: {carpeta_trabajo}")
    
    
    
    try:
        # 1. Cargar esquema
        esquema_path = resolve_structured_path("assets/esquemas/esquema_nominas")
        
        if not esquema_path.exists():
            raise FileNotFoundError(f"Esquema no encontrado: {esquema_path}")
        
        esquema = load_structured_data(esquema_path, prefer_resource_path=False)
        
        print(f"   ✓ Esquema cargado: v{esquema['metadata']['version']}")
        
        # 2. Leer datos Silver
        import polars as pl
        df_silver = pl.read_parquet(ruta_parquet_silver)
        
        print(f"   ✓ Silver cargado: {len(df_silver):,} registros")
        
        # 3. Transformar a Gold
        df_gold = seleccionar_y_convertir_columnas(df_silver, esquema)
        
        print(f"   ✓ Transformaciones aplicadas: {len(df_gold):,} registros")
        
        # 4. Guardar Gold con estructura simplificada
        carpeta_silver = ruta_parquet_silver.parent
        rutas_gold = guardar_resultados(
            df_gold,
            carpeta_silver,
            export_excel=export_excel_gold,
        )
        
        print(f"   ✓ Gold guardado exitosamente")
        print(f"     • Estructura: gold/")
        print(f"       └── actual/    (archivos actuales)")
        
        return {
            'success': True,
            'parquet': rutas_gold['parquet'],
            'excel': rutas_gold['excel'],
            'carpeta_actual': rutas_gold['carpeta_actual'],
            'registros': len(df_gold),
            'columnas': len(df_gold.columns)
        }
        
    except Exception as e:
        print(f"   ✗ Error: {e}")
        import traceback
        traceback.print_exc()
        raise


# ============================================================================
# FUNCIÓN MAIN PARA EJECUCIÓN STANDALONE
# ============================================================================

def main():
    """Función main para ejecución standalone con diálogos de archivo"""
    from tkinter import Tk, filedialog
    
    print("=" * 70)
    print("TRANSFORMACIÓN SILVER → GOLD - REPORTES DE PLANILLA")
    print("=" * 70)
    print()
    
    # Seleccionar parquet silver
    print("🔍 Seleccione el archivo Parquet Silver...")
    
    root = Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    
    ruta_parquet = filedialog.askopenfilename(
        title="Seleccione el archivo Parquet Silver",
        filetypes=[("Parquet files", "*.parquet"), ("All files", "*.*")]
    )
    root.destroy()
    
    if not ruta_parquet:
        print("❌ No se seleccionó archivo parquet. Operación cancelada.")
        return
    
    print(f"✓ Archivo seleccionado: {Path(ruta_parquet).name}")
    print()
    
    # Buscar carpeta de esquemas - buscar en directorio actual y niveles superiores
    carpeta_actual = Path.cwd()
    carpeta_esquemas = None
    
    # Buscar en el directorio actual y hasta 3 niveles arriba
    for _ in range(4):
        posible_esquemas = carpeta_actual / "esquemas"
        if posible_esquemas.exists() and posible_esquemas.is_dir():
            carpeta_esquemas = posible_esquemas
            break
        carpeta_actual = carpeta_actual.parent
    
    # Si no se encontró, crear en el directorio donde se ejecuta el script
    if carpeta_esquemas is None:
        carpeta_proyecto = Path.cwd()
        carpeta_esquemas = carpeta_proyecto / "esquemas"
        
        if not carpeta_esquemas.exists():
            print(f"⚠️  No se encontró la carpeta 'esquemas'")
            print(f"📁 Creando carpeta 'esquemas' en: {carpeta_proyecto}")
            carpeta_esquemas.mkdir(exist_ok=True)
            print(f"✓ Carpeta creada: {carpeta_esquemas}")
            print()
            print("❌ Por favor, coloca los archivos de esquema (.yaml/.yml) en esta carpeta y ejecuta nuevamente.")
            return

    # Listar archivos de esquema en la carpeta de esquemas
    esquemas_disponibles = list_structured_files(carpeta_esquemas)

    if not esquemas_disponibles:
        print(f"❌ No se encontraron archivos de esquema en: {carpeta_esquemas}")
        print(f"   Por favor, coloca los archivos de esquemas (.yaml/.yml) en esta carpeta.")
        return
    
    print(f"📁 Carpeta de esquemas: {carpeta_esquemas}")
    print(f"✓ Esquemas disponibles:")
    for i, esquema in enumerate(esquemas_disponibles, 1):
        print(f"   {i}. {esquema.name}")
    print()
    
    # Seleccionar schema YAML de la carpeta de esquemas
    print("🔍 Seleccione el archivo del esquema Gold (YAML)...")
    
    root = Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    
    ruta_schema = filedialog.askopenfilename(
        title="Seleccione el esquema Gold (YAML)",
        initialdir=str(carpeta_esquemas),
        filetypes=structured_filetypes()
    )
    root.destroy()
    
    if not ruta_schema:
        print("❌ No se seleccionó archivo de esquema. Operación cancelada.")
        return
    
    print(f"✓ Esquema seleccionado: {Path(ruta_schema).name}")
    print()
    
    # Cargar datos
    print("📊 Cargando datos silver...")
    inicio = datetime.now()
    df = pl.read_parquet(ruta_parquet)
    print(f"✓ Datos cargados: {df.shape[0]:,} filas × {df.shape[1]} columnas")
    
    # Cargar schema
    print("\n📋 Cargando esquema gold...")
    schema = load_structured_data(ruta_schema, prefer_resource_path=False)
    print(f"✓ Schema versión {schema['metadata']['version']} cargado")
    print(f"✓ Columnas esperadas: {len(schema['schema'])}")
    
    # Aplicar transformaciones
    print("\n" + "=" * 70)
    print("APLICANDO TRANSFORMACIONES GOLD")
    print("=" * 70)
    
    try:
        df_gold = seleccionar_y_convertir_columnas(df, schema)
        
        print(f"\n✓ Transformaciones aplicadas exitosamente")
        print(f"  - Columnas finales: {df_gold.shape[1]}")
        print(f"  - Registros: {df_gold.shape[0]:,}")
    except Exception as e:
        print(f"\n❌ Error al aplicar transformaciones: {e}")
        import traceback
        traceback.print_exc()
        return
    
    # Guardar archivos
    carpeta_silver = Path(ruta_parquet).parent
    rutas = guardar_resultados(df_gold, carpeta_silver)
    
    # Resumen final
    duracion = (datetime.now() - inicio).total_seconds()
    print("\n" + "=" * 70)
    print("✅ PROCESO COMPLETADO EXITOSAMENTE")
    print("=" * 70)
    print(f"⏱️  Tiempo de ejecución: {duracion:.2f} segundos")
    print(f"📊 Registros procesados: {df_gold.shape[0]:,}")
    print(f"📋 Schema utilizado: {Path(ruta_schema).name}")
    print(f"\n📁 Estructura de carpetas Gold (SIMPLIFICADA):")
    print(f"   gold/")
    print(f"   └── actual/        (Power BI apunta aquí)")
    print(f"       ├── Planilla_Metso_Consolidado.parquet")
    if rutas['excel'] is not None:
        print(f"       └── Planilla_Metso_Consolidado.xlsx")
    
    print("\n💡 Estructura simplificada - sin carpeta 'nomina' intermedia")
    print("💡 Los archivos en actual/ se sobreescriben en cada ejecución")
    print("=" * 70)


if __name__ == "__main__":
    main()
