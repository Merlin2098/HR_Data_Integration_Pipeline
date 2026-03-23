"""
Script: step1_consolidar_licencias.py
Descripción: Procesa archivo Excel CONTROL_DE_LICENCIAS.xlsx
             - Hoja LICENCIA CON GOCE: headers en fila 2, datos desde fila 3
             - Hoja LICENCIA SIN GOCE: headers en fila 2, datos desde fila 3
             
Arquitectura:
- Bronze: Excel con licencias con goce y sin goce
- Silver: Parquet consolidado con versionamiento

Salida: 
    - /actual/Licencias_Consolidadas.parquet (sin timestamp para Power BI)
    - /historico/Licencias_Consolidadas_YYYYMMDD_HHMMSS.parquet (versionado)

Autor: Richi via Claude
Fecha: 26.01.2026
"""

import polars as pl
import openpyxl
from pathlib import Path
from datetime import datetime
import time
import sys
from tkinter import Tk, filedialog

from src.utils.structured_config import load_structured_data, resolve_structured_path


def seleccionar_archivo_excel() -> Path | None:
    """Abre diálogo para seleccionar archivo Excel"""
    root = Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    
    archivo = filedialog.askopenfilename(
        title="Seleccionar archivo CONTROL_DE_LICENCIAS.xlsx (Bronze)",
        filetypes=[("Excel files", "*.xlsx *.xlsm *.xls"), ("All files", "*.*")]
    )
    
    root.destroy()
    
    return Path(archivo) if archivo else None


def cargar_esquema(ruta_esquema: Path) -> dict:
    """Carga el esquema de validación YAML."""
    try:
        return load_structured_data(ruta_esquema, prefer_resource_path=False)
    except Exception as e:
        print(f"   ✗ Error al cargar esquema: {e}")
        raise


def validar_esquema(df: pl.DataFrame, esquema: dict) -> tuple[bool, list[str]]:
    """
    Valida que el DataFrame cumpla con el esquema definido.
    Retorna (es_valido, lista_errores)
    """
    errores = []
    columnas_df = set(df.columns)
    
    # Validar columnas obligatorias
    for col_def in esquema["columns"]:
        nombre_col = col_def["name"]
        
        if nombre_col not in columnas_df:
            errores.append(f"Columna obligatoria faltante: {nombre_col}")
            continue
        
        # Validar nullabilidad
        if not col_def["nullable"]:
            null_count = df[nombre_col].null_count()
            if null_count > 0:
                errores.append(
                    f"Columna '{nombre_col}' contiene {null_count} valores nulos (no permitido)"
                )
        
        # Validar tipo de dato
        tipo_esperado = col_def["type"]
        tipo_actual = str(df[nombre_col].dtype)
        
        # Mapeo flexible de tipos
        mapeo_tipos = {
            "string": ["Utf8", "String"],
            "integer": ["Int64", "Int32", "Int16", "Int8"],
            "date": ["Date", "Datetime"],
        }
        
        if tipo_esperado in mapeo_tipos:
            if not any(t in tipo_actual for t in mapeo_tipos[tipo_esperado]):
                errores.append(
                    f"Columna '{nombre_col}': tipo esperado {tipo_esperado}, encontrado {tipo_actual}"
                )
        
        # Validaciones adicionales
        if "validations" in col_def:
            validaciones = col_def["validations"]
            
            # Validar rango numérico
            if "min" in validaciones and tipo_esperado == "integer":
                min_val = df[nombre_col].min()
                if min_val and min_val < validaciones["min"]:
                    errores.append(
                        f"Columna '{nombre_col}': valor mínimo {min_val} menor que {validaciones['min']}"
                    )
            
            if "max" in validaciones and tipo_esperado == "integer":
                max_val = df[nombre_col].max()
                if max_val and max_val > validaciones["max"]:
                    errores.append(
                        f"Columna '{nombre_col}': valor máximo {max_val} mayor que {validaciones['max']}"
                    )
            
            # Validar valores permitidos
            if "allowed_values" in validaciones:
                valores_unicos = df[nombre_col].unique().to_list()
                valores_invalidos = [
                    v for v in valores_unicos 
                    if v not in validaciones["allowed_values"] and v is not None
                ]
                if valores_invalidos:
                    errores.append(
                        f"Columna '{nombre_col}': valores no permitidos: {valores_invalidos}"
                    )
    
    return len(errores) == 0, errores


def leer_hoja_excel(
    ruta_archivo: Path, 
    nombre_hoja: str, 
    tipo_licencia: str,
    esquema: dict
) -> pl.DataFrame | None:
    """
    Lee una hoja del archivo Excel desde la fila 2 (headers).
    Aplica filtros de seguridad y normalización.
    
    Args:
        ruta_archivo: Path al archivo Excel
        nombre_hoja: Nombre de la hoja a procesar
        tipo_licencia: "CON_GOCE" o "SIN_GOCE"
        esquema: Diccionario con el esquema de validación
        
    Returns:
        DataFrame con los datos procesados o None si hay error
    """
    print(f"\n   → Procesando hoja: {nombre_hoja}")
    
    try:
        wb = openpyxl.load_workbook(ruta_archivo, read_only=True, data_only=True)
        
        if nombre_hoja not in wb.sheetnames:
            print(f"   ⚠️  Hoja '{nombre_hoja}' no encontrada, omitiendo...")
            wb.close()
            return None
        
        ws = wb[nombre_hoja]
        
        # Leer headers desde fila 2
        headers_raw = [cell.value for cell in ws[2]]
        
        # Filtrar headers no vacíos
        headers = []
        indices_validos = []
        for idx, header in enumerate(headers_raw):
            if header is not None and str(header).strip() != "":
                headers.append(str(header).strip())
                indices_validos.append(idx)
        
        if not headers:
            print(f"   ⚠️  No se encontraron headers válidos en '{nombre_hoja}'")
            wb.close()
            return None
        
        print(f"   ✓ Headers válidos encontrados: {len(headers)}")
        
        # Leer datos desde fila 3 en adelante
        datos = []
        filas_procesadas = 0
        filas_filtradas = 0
        
        for fila in ws.iter_rows(min_row=3, values_only=True):
            # FILTRO DE SEGURIDAD: Columna A (DNI) no puede ser nulo/vacío
            dni_valor = fila[0] if len(fila) > 0 else None
            
            if dni_valor is None or str(dni_valor).strip() == "":
                filas_filtradas += 1
                continue
            
            # Extraer solo columnas con headers válidos
            fila_filtrada = [fila[idx] if idx < len(fila) else None for idx in indices_validos]
            datos.append(fila_filtrada)
            filas_procesadas += 1
        
        wb.close()
        
        if filas_filtradas > 0:
            print(f"   ⚠️  Filas filtradas por DNI vacío: {filas_filtradas}")
        
        if not datos:
            print(f"   ⚠️  No se encontraron datos válidos en '{nombre_hoja}'")
            return None
        
        # Crear DataFrame
        df = pl.DataFrame(datos, schema=headers, orient="row")
        
        print(f"   ✓ Filas leídas: {filas_procesadas}")
        
        # MAPEO DE COLUMNAS: Leer desde esquema de configuración
        mapeo_columnas = esquema.get("column_mapping", {})
        
        # Aplicar renombrado de columnas si existen
        for col_original, col_nueva in mapeo_columnas.items():
            if col_original in df.columns:
                df = df.rename({col_original: col_nueva})
        
        # Normalizar DNI/CEX a string (quitar .0 si viene como float)
        if "DNI/CEX" in df.columns:
            df = df.with_columns(
                pl.col("DNI/CEX").cast(pl.Utf8).str.replace(r"\.0$", "").alias("DNI/CEX")
            )
        
        # Mantener PERIODO como string (formato YYYY-MM)
        if "PERIODO" in df.columns:
            df = df.with_columns(
                pl.col("PERIODO").cast(pl.Utf8).alias("PERIODO")
            )
        
        # Validar MOTIVO no vacío
        if "MOTIVO" in df.columns:
            df = df.filter(
                pl.col("MOTIVO").is_not_null() &
                (pl.col("MOTIVO").cast(pl.Utf8).str.strip_chars() != "")
            )
        
        # Agregar columna TIPO_LICENCIA
        df = df.with_columns(
            pl.lit(tipo_licencia).alias("TIPO_LICENCIA")
        )
        
        # Aplicar filtros de seguridad adicionales
        df = df.filter(
            # DNI/CEX no vacío después de strip
            (pl.col("DNI/CEX").str.strip_chars() != "") &
            # PERIODO no vacío
            (pl.col("PERIODO").str.strip_chars() != "")
        )
        
        print(f"   ✓ Filas después de filtros: {len(df)}")
        
        return df
        
    except Exception as e:
        print(f"   ✗ Error procesando '{nombre_hoja}': {e}")
        return None


def consolidar_y_exportar(
    df_consolidado: pl.DataFrame, 
    carpeta_trabajo: Path,
    registros_con_goce: int,
    registros_sin_goce: int
):
    """
    Exporta el DataFrame consolidado a Parquet.
    Se sobreescribe en cada ejecución (sin versionamiento).
    
    Args:
        df_consolidado: DataFrame a exportar
        carpeta_trabajo: Path de la carpeta de trabajo
        registros_con_goce: Cantidad de registros con goce
        registros_sin_goce: Cantidad de registros sin goce
    """
    # Crear carpeta silver (sin subcarpeta licencias)
    carpeta_silver = carpeta_trabajo / "silver"
    carpeta_silver.mkdir(parents=True, exist_ok=True)
    
    print(f"\n[2/2] Guardando resultados en capa Silver...")
    print(f"  📁 Carpeta: {carpeta_silver}")
    
    nombre_base = "licencias_consolidadas"
    
    # Guardar Parquet (se sobreescribe)
    print(f"\n  - Guardando parquet...", end='', flush=True)
    ruta_parquet = carpeta_silver / f"{nombre_base}.parquet"
    df_consolidado.write_parquet(ruta_parquet, compression="snappy")
    print(f" ✓")
    print(f"    Ubicación: {ruta_parquet.name}")
    
    # Resumen de distribución
    print(f"\n📊 Distribución de registros:")
    print(f"  - CON GOCE:  {registros_con_goce:,}")
    print(f"  - SIN GOCE:  {registros_sin_goce:,}")
    print(f"  - TOTAL:     {len(df_consolidado):,}")


def main():
    """Función principal de procesamiento"""
    print("=" * 80)
    print(" CONSOLIDADOR DE LICENCIAS - CAPA SILVER ".center(80, "="))
    print("=" * 80)
    
    # 1. Seleccionar archivo
    print("\n[PASO 1] Selecciona el archivo CONTROL_DE_LICENCIAS.xlsx (Bronze)...")
    archivo_bronze = seleccionar_archivo_excel()
    
    if not archivo_bronze:
        print("✗ No se seleccionó ningún archivo. Proceso cancelado.")
        return
    
    # Iniciar cronómetro después de la selección
    tiempo_inicio = time.time()
    
    print(f"✓ Archivo seleccionado: {archivo_bronze.name}")
    carpeta_trabajo = archivo_bronze.parent
    
    # 2. Cargar esquema
    print("\n" + "=" * 80)
    print(" PROCESAMIENTO ".center(80, "="))
    print("=" * 80)
    print(f"\n[1/2] Cargando esquema de validación...")
    
    try:
        ruta_esquema = resolve_structured_path("assets/esquemas/esquema_licencias")
        esquema = cargar_esquema(ruta_esquema)
        print(f"   ✓ Esquema cargado: {ruta_esquema.name}")
    except Exception as e:
        print(f"✗ Error al cargar esquema: {e}")
        return
    
    # 3. Procesar hojas
    print(f"\n   Extrayendo y consolidando datos...")
    
    try:
        # Hoja 1: CON GOCE
        df_con_goce = leer_hoja_excel(
            archivo_bronze, 
            "LICENCIA CON GOCE", 
            "CON_GOCE",
            esquema
        )
        
        # Hoja 2: SIN GOCE
        df_sin_goce = leer_hoja_excel(
            archivo_bronze, 
            "LICENCIA SIN GOCE", 
            "SIN_GOCE",
            esquema
        )
        
        # Validar que al menos una hoja tenga datos
        if df_con_goce is None and df_sin_goce is None:
            print("\n✗ No se pudo procesar ninguna hoja válida")
            return
        
        # Consolidar DataFrames
        dfs_validos = [df for df in [df_con_goce, df_sin_goce] if df is not None]
        df_consolidado = pl.concat(dfs_validos, how="diagonal")
        
        # Contar registros por tipo
        registros_con_goce = len(df_con_goce) if df_con_goce is not None else 0
        registros_sin_goce = len(df_sin_goce) if df_sin_goce is not None else 0
        
        print(f"\n   ✓ Total registros consolidados: {len(df_consolidado):,}")
        
        # Seleccionar solo columnas del esquema
        columnas_esquema = [col["name"] for col in esquema["columns"]]
        columnas_disponibles = [col for col in columnas_esquema if col in df_consolidado.columns]
        df_consolidado = df_consolidado.select(columnas_disponibles)
        
        # Validar esquema
        print(f"\n   Validando esquema...")
        es_valido, errores = validar_esquema(df_consolidado, esquema)
        
        if not es_valido:
            print("   ✗ Errores de validación:")
            for error in errores:
                print(f"     • {error}")
            return
        
        print(f"   ✓ Esquema válido")
        
        # Exportar resultados
        consolidar_y_exportar(
            df_consolidado, 
            carpeta_trabajo,
            registros_con_goce,
            registros_sin_goce
        )
        
        # Calcular tiempo total
        tiempo_total = time.time() - tiempo_inicio
        
        # 4. Resumen final
        print("\n" + "=" * 80)
        print(" RESUMEN ".center(80, "="))
        print("=" * 80)
        
        print(f"\n✓ Procesamiento completado exitosamente")
        
        print(f"\n📂 Archivos generados:")
        print(f"  - licencias_consolidadas.parquet")
        
        print(f"\n⏱️  Tiempo de ejecución: {tiempo_total:.2f}s")
        
        print("\n💡 Los archivos se sobreescriben en cada ejecución (sin historial)")
        
        print("\n" + "=" * 80)
        
    except Exception as e:
        print(f"\n✗ Error durante el procesamiento: {str(e)}")
        import traceback
        traceback.print_exc()
        return


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"\n✗ Error fatal: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

def procesar_sin_gui(ruta_archivo: Path, carpeta_salida: Path) -> dict:
    """
    Procesa licencias sin interfaz gráfica (modo headless)
    Usado por el pipeline executor
    
    Args:
        ruta_archivo: Path al archivo CONTROL_DE_LICENCIAS.xlsx
        carpeta_salida: Path a la carpeta /silver/ donde guardar el parquet
        
    Returns:
        dict con resultados del procesamiento
    """
    print(f"\n🔄 Procesando licencias (modo headless)...")
    print(f"   Archivo: {ruta_archivo.name}")
    print(f"   Salida: {carpeta_salida}")
    
    try:
        # Cargar esquema
        ruta_esquema = resolve_structured_path("assets/esquemas/esquema_licencias")
        esquema = cargar_esquema(ruta_esquema)
        
        # Procesar ambas hojas
        df_con_goce = leer_hoja_excel(
            ruta_archivo, 
            "LICENCIA CON GOCE", 
            "CON_GOCE",
            esquema
        )
        
        df_sin_goce = leer_hoja_excel(
            ruta_archivo, 
            "LICENCIA SIN GOCE", 
            "SIN_GOCE",
            esquema
        )
        
        # Validar que al menos una hoja tenga datos
        if df_con_goce is None and df_sin_goce is None:
            raise ValueError("No se pudo procesar ninguna hoja válida")
        
        # Consolidar DataFrames
        dfs_validos = [df for df in [df_con_goce, df_sin_goce] if df is not None]
        df_consolidado = pl.concat(dfs_validos, how="diagonal")
        
        registros_con_goce = len(df_con_goce) if df_con_goce is not None else 0
        registros_sin_goce = len(df_sin_goce) if df_sin_goce is not None else 0
        
        print(f"   ✓ Registros consolidados: {len(df_consolidado):,}")
        
        # Seleccionar solo columnas del esquema
        columnas_esquema = [col["name"] for col in esquema["columns"]]
        columnas_disponibles = [col for col in columnas_esquema if col in df_consolidado.columns]
        df_consolidado = df_consolidado.select(columnas_disponibles)
        
        # Validar esquema
        es_valido, errores = validar_esquema(df_consolidado, esquema)
        
        if not es_valido:
            print("   ✗ Errores de validación:")
            for error in errores:
                print(f"     • {error}")
            raise ValueError("Validación de esquema falló")
        
        # Guardar parquet
        carpeta_salida.mkdir(parents=True, exist_ok=True)
        ruta_parquet = carpeta_salida / "licencias_consolidadas.parquet"
        
        df_consolidado.write_parquet(ruta_parquet, compression="snappy")
        
        print(f"   ✓ Parquet guardado: {ruta_parquet.name}")
        
        return {
            'success': True,
            'parquet': ruta_parquet,
            'registros': len(df_consolidado),
            'registros_con_goce': registros_con_goce,
            'registros_sin_goce': registros_sin_goce
        }
        
    except Exception as e:
        print(f"   ✗ Error: {e}")
        raise
