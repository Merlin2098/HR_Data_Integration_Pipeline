"""
Script de transformación Silver → Gold para reportes de planilla - Régimen Minero
Guarda salida Gold únicamente en carpeta actual/
"""

import polars as pl
from pathlib import Path
from tkinter import Tk, filedialog
from datetime import datetime

from src.utils.structured_config import (
    list_structured_files,
    load_structured_data,
    structured_filetypes,
)
from src.utils.month_name import add_month_name_column


def seleccionar_archivo(titulo, tipos):
    """Abre un diálogo para seleccionar archivo"""
    root = Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    
    archivo = filedialog.askopenfilename(
        title=titulo,
        filetypes=tipos
    )
    root.destroy()
    return archivo


def aplicar_transformaciones_gold(df, schema):
    """Aplica transformaciones según el esquema gold"""
    
    # NO normalizar columnas - mantener nombres originales del parquet
    print(f"\n📋 Seleccionando columnas del schema...")
    print("-" * 70)
    
    # Seleccionar solo las columnas que existen en el schema
    columnas_schema = list(schema["schema"].keys())
    columnas_disponibles = [col for col in columnas_schema if col in df.columns]
    columnas_faltantes = [col for col in columnas_schema if col not in df.columns]
    
    for col in columnas_disponibles:
        print(f"✓ {col}")
    
    if columnas_faltantes:
        print(f"\n⚠️  Columnas del schema no encontradas en el parquet:")
        for col in columnas_faltantes:
            print(f"  ✗ {col}")
    
    print("-" * 70)
    print(f"Total: {len(columnas_disponibles)} de {len(columnas_schema)} columnas")
    print()
    
    # Seleccionar columnas
    df = df.select(columnas_disponibles)
    
    # Aplicar tipos de datos según schema
    print("🔄 Aplicando tipos de datos...")
    for col_name in columnas_disponibles:
        col_spec = schema["schema"][col_name]
        col_type = col_spec["type"]
        
        try:
            if col_type == "string":
                df = df.with_columns(pl.col(col_name).cast(pl.Utf8, strict=False))
            elif col_type == "integer":
                df = df.with_columns(pl.col(col_name).cast(pl.Int64, strict=False))
            elif col_type == "float":
                df = df.with_columns(pl.col(col_name).cast(pl.Float64, strict=False))
            elif col_type == "date":
                # Manejar fechas que vienen como datetime string
                df = df.with_columns(
                    pl.col(col_name).str.strptime(pl.Datetime, "%Y-%m-%d %H:%M:%S%.f", strict=False).cast(pl.Date)
                )
        except Exception as e:
            print(f"  ⚠️  Error al convertir columna {col_name}: {e}")
    
    # Aplicar string trimming
    print("✂️  Limpiando espacios en strings...")
    for col in df.columns:
        if df[col].dtype == pl.Utf8:
            df = df.with_columns(pl.col(col).str.strip_chars())
    
    return df


def agregar_nombre_mes(df):
    """Agrega columna con nombre del mes en inglés."""
    
    print("📅 Agregando nombre del mes...")

    df = add_month_name_column(df, default_invalid=None)
    
    print(f"✓ Columna 'NOMBRE_MES' agregada después de 'MES'")
    
    return df


def generar_excel_visualizacion(df, ruta_salida):
    """Genera Excel con formato para visualización"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    print("📝 Generando Excel con formato...")
    
    # Convertir a pandas para usar openpyxl
    df_pandas = df.to_pandas()
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Planilla"
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Escribir encabezados
    for col_idx, col_name in enumerate(df_pandas.columns, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = col_name
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # Escribir datos
    for row_idx, row_data in enumerate(df_pandas.values, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = border
            
            # Alineación según tipo de dato
            if isinstance(value, (int, float)):
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="left")
    
    # Ajustar anchos de columna
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Congelar primera fila
    ws.freeze_panes = "A2"
    
    # Guardar
    wb.save(ruta_salida)


def gestionar_versionamiento_gold(carpeta_base):
    """
    Prepara la salida en la capa Gold:
    - Crea carpeta gold/actual/
    
    Args:
        carpeta_base: Path de la carpeta base del proyecto
        
    Returns:
        Path: carpeta_actual
    """
    # Crear estructura de carpetas
    carpeta_gold = Path(carpeta_base) / "gold"
    carpeta_actual = carpeta_gold / "actual"
    
    # Crear carpetas si no existen
    carpeta_actual.mkdir(parents=True, exist_ok=True)
    
    return carpeta_actual


def main():
    print("=" * 70)
    print("TRANSFORMACIÓN SILVER → GOLD - RÉGIMEN MINERO")
    print("=" * 70)
    print()
    
    # Seleccionar parquet silver
    print("🔍 Seleccione el archivo Parquet Silver...")
    ruta_parquet = seleccionar_archivo(
        "Seleccione el archivo Parquet Silver - Régimen Minero",
        [("Parquet files", "*.parquet"), ("All files", "*.*")]
    )
    
    if not ruta_parquet:
        print("❌ No se seleccionó archivo parquet. Operación cancelada.")
        return
    
    print(f"✓ Archivo seleccionado: {Path(ruta_parquet).name}")
    print()
    
    # Buscar carpeta de esquemas - buscar en directorio actual y niveles superiores
    carpeta_actual = Path.cwd()
    carpeta_esquemas = None
    
    # Buscar en el directorio actual y hasta 3 niveles arriba
    for _ in range(4):
        posible_esquemas = carpeta_actual / "esquemas"
        if posible_esquemas.exists() and posible_esquemas.is_dir():
            carpeta_esquemas = posible_esquemas
            break
        carpeta_actual = carpeta_actual.parent
    
    # Si no se encontró, crear en el directorio donde se ejecuta el script
    if carpeta_esquemas is None:
        carpeta_proyecto = Path.cwd()
        carpeta_esquemas = carpeta_proyecto / "esquemas"
        
        if not carpeta_esquemas.exists():
            print(f"⚠️  No se encontró la carpeta 'esquemas'")
            print(f"📁 Creando carpeta 'esquemas' en: {carpeta_proyecto}")
            carpeta_esquemas.mkdir(exist_ok=True)
            print(f"✓ Carpeta creada: {carpeta_esquemas}")
            print()
            print("❌ Por favor, coloca los archivos de esquema (.yaml/.yml) en esta carpeta y ejecuta nuevamente.")
            return

    # Listar archivos de esquema en la carpeta de esquemas
    esquemas_disponibles = list_structured_files(carpeta_esquemas)

    if not esquemas_disponibles:
        print(f"❌ No se encontraron archivos de esquema en: {carpeta_esquemas}")
        print(f"   Por favor, coloca los archivos de esquemas (.yaml/.yml) en esta carpeta.")
        return
    
    print(f"📁 Carpeta de esquemas: {carpeta_esquemas}")
    print(f"✓ Esquemas disponibles:")
    for i, esquema in enumerate(esquemas_disponibles, 1):
        print(f"   {i}. {esquema.name}")
    print()
    
    # Seleccionar schema YAML de la carpeta de esquemas
    print("🔍 Seleccione el archivo del esquema Gold (Régimen Minero)...")
    
    # Cambiar directorio inicial del diálogo a la carpeta de esquemas
    root = Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    
    ruta_schema = filedialog.askopenfilename(
        title="Seleccione el esquema Gold - Régimen Minero (YAML)",
        initialdir=str(carpeta_esquemas),
        filetypes=structured_filetypes()
    )
    root.destroy()
    
    if not ruta_schema:
        print("❌ No se seleccionó archivo schema. Operación cancelada.")
        return
    
    print(f"✓ Esquema seleccionado: {Path(ruta_schema).name}")
    print()
    
    # Cargar datos
    print("📊 Cargando datos silver...")
    inicio = datetime.now()
    df = pl.read_parquet(ruta_parquet)
    print(f"✓ Datos cargados: {df.shape[0]:,} filas × {df.shape[1]} columnas")
    
    # Cargar schema
    print("\n📋 Cargando esquema gold...")
    schema = load_structured_data(ruta_schema, prefer_resource_path=False)
    print(f"✓ Schema versión {schema['metadata']['version']} cargado")
    print(f"✓ Columnas esperadas: {len(schema['schema'])}")
    
    # Aplicar transformaciones
    print("\n" + "=" * 70)
    print("APLICANDO TRANSFORMACIONES GOLD")
    print("=" * 70)
    
    try:
        # Aplicar transformaciones según esquema
        df_gold = aplicar_transformaciones_gold(df, schema)
        
        # Agregar columna NOMBRE_MES
        df_gold = agregar_nombre_mes(df_gold)
        
        print(f"\n✓ Transformaciones aplicadas exitosamente")
        print(f"  - Columnas finales: {df_gold.shape[1]}")
        print(f"  - Registros: {df_gold.shape[0]:,}")
    except Exception as e:
        print(f"\n❌ Error al aplicar transformaciones: {e}")
        import traceback
        traceback.print_exc()
        return
    
    # Preparar carpeta gold
    carpeta_base = Path(ruta_parquet).parent.parent  # Subir desde silver/ a carpeta base
    carpeta_actual = gestionar_versionamiento_gold(carpeta_base)
    
    # Rutas de salida en actual/ (sin timestamp)
    ruta_parquet_gold_actual = carpeta_actual / "Planilla Metso - Regimen Minero.parquet"
    ruta_excel_gold_actual = carpeta_actual / "Planilla Metso - Regimen Minero.xlsx"
    
    # Guardar archivos
    print("\n💾 Guardando archivos en capa Gold...")
    print(f"  📁 Carpeta actual: {carpeta_actual}")
    print("-" * 70)
    
    # Guardar en actual/
    df_gold.write_parquet(ruta_parquet_gold_actual)
    print(f"✓ Parquet gold (actual): {ruta_parquet_gold_actual.name}")
    
    # Generar Excel de visualización en actual/
    try:
        generar_excel_visualizacion(df_gold, ruta_excel_gold_actual)
        print(f"✓ Excel gold (actual): {ruta_excel_gold_actual.name}")
    except Exception as e:
        print(f"⚠️  Error al generar Excel en actual/: {e}")
    
    # Resumen final
    duracion = (datetime.now() - inicio).total_seconds()
    print("\n" + "=" * 70)
    print("✅ PROCESO COMPLETADO EXITOSAMENTE")
    print("=" * 70)
    print(f"⏱️  Tiempo de ejecución: {duracion:.2f} segundos")
    print(f"📊 Registros procesados: {df_gold.shape[0]:,}")
    print(f"📋 Schema utilizado: {Path(ruta_schema).name}")
    print(f"\n📁 Estructura de carpetas Gold:")
    print(f"   {carpeta_base / 'gold'}/")
    print(f"   └── actual/        (Power BI apunta aquí - se sobreescribe)")
    print(f"       ├── Planilla Metso - Regimen Minero.parquet")
    print(f"       └── Planilla Metso - Regimen Minero.xlsx")
    
    print("\n💡 Los archivos en actual/ se sobreescriben en cada ejecución")
    print("=" * 70)


if __name__ == "__main__":
    main()
